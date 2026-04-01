"""
scrape_all_brands.py
Multi-brand Indian bicycle price scraper.
Produces:  all_bike_prices.xlsx  (one sheet per brand)
"""

import re, time, json, sys
from copy import deepcopy
from pathlib import Path
from urllib.parse import urlparse, unquote, urljoin
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ─────────────────────────── headers ──────────────────────────────────────────
HDR = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-IN,en;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

COLS = ["Model name", "Segment", "Listed price", "Discounted price"]

# FX cache for converting non-INR prices to INR.
FX_CACHE = {"ts": 0.0, "rates": {"INR": 1.0}}
FX_FALLBACK = {
    "INR": 1.0,
    "USD": 83.0,
    "EUR": 90.0,
    "GBP": 105.0,
    "CHF": 94.0,
    "AUD": 54.0,
    "CAD": 61.0,
    "SGD": 62.0,
    "AED": 22.6,
    "PHP": 1.45,
    "MYR": 17.8,
    "IDR": 0.0052,
    "THB": 2.3,
    "JPY": 0.55,
    "CNY": 11.5,
    "KRW": 0.062,
}

DEFAULT_CLIENT_BRANDS = [
    ("Trek", "https://www.trekbikes.com/in/en_IN/"),
    ("Giant", "https://giantindia.com/"),
    ("Scott", "https://sportnetwork.in/"),
    ("Cannondale", "https://www.cannondale.com/en"),
    ("Merida", "https://www.merida-bikes.com/en"),
    ("Specialized", "https://www.specialized.com/us/en"),
    ("Polygon", "https://www.polygonbikes.com/"),
    ("Avanti", "https://sportnetwork.in/"),
    ("Bergamont", "https://sportnetwork.in/"),
    ("BMC", "https://bmc-switzerland.com/"),
    ("Cervelo", "https://www.cervelo.com/en-us"),
    ("Pinarello", "https://pinarello.com/usa/en"),
    ("Look", "https://www.lookcycle.com/qe-en/"),
    ("Felt", "https://feltbicycles.com/collections/all-bikes"),
    ("Fuji", "https://www.fujibikes.com/collections/all-bikes"),
    ("Java", "https://javabikesph.com/"),
    ("Jamis", "https://www.jamisbikes.com/int/"),
    ("Surly", "https://surlybikes.com/"),
    ("Basso", "https://bassobikes.com/en"),
    ("Kona", "https://konaworld.com/"),
    ("Marin", "https://marinbikes.com/"),
    ("Orbea", "https://www.orbea.com/in-en/"),
    ("Ridley", "https://www.ridley-bikes.com/en_IN"),
    ("Factor", "https://factorbikes.com/"),
    ("Canyon", "https://www.canyon.com/en-in/"),
    ("Cube", "https://www.cube.eu/"),
]

# ─────────────────────────── brand registry ───────────────────────────────────
# platform: shopify | occ | woocommerce | sfcc | playwright | decathlon | models_generic | dynamic_models | unavailable
BRANDS = [
    # ── Shopify ────────────────────────────────────────────────────────────────
    {
        "name": "Firefox Bikes",
        "platform": "shopify",
        "base_url": "https://www.firefoxlife.com",
    },
    {
        "name": "Cradiac",
        "platform": "shopify",
        "base_url": "https://www.cradiac.com",
    },
    {
        "name": "Caya Bikes",
        "platform": "shopify",
        "base_url": "https://www.cayabikes.com",
    },
    {
        "name": "Leader Bicycles",
        "platform": "shopify",
        "base_url": "https://www.leaderbicycles.com",
    },
    {
        "name": "EMotorad",
        "platform": "unavailable",
        "note": "Shopify store is password-protected (returns same 8249-byte lock page for all URLs)",
    },
    # ── SAP Commerce Cloud (OCC) -> Swapped to BOTS Shopify for Indian Pricing  ──
    {
        "name": "Trek",
        "platform": "shopify",
        "base_url": "https://bumsonthesaddle.com/collections/trek-bikes",
    },
    # ── WooCommerce (HTML scrape of listing pages) ────────────────────────────
    {
        "name": "Hero Cycles",
        "platform": "woocommerce",
        "base_url": "https://www.herocycles.com",
        "catalog_paths": ["/bikes/"],
    },
    {
        "name": "Avon Cycles",
        "platform": "woocommerce",
        "base_url": "https://www.avoncycles.com",
        "catalog_paths": ["/shop/"],
        "card_sel":   "div.item",
        "price_sels": ["span.new-price"],
        "mrp_sels":   ["span.old-price"],
        "name_sels":  ["div.product-name a", "h2", "h3", ".name"],
        "link_sel":   "a[href]",
    },
    # ── Hero Lectro — SFCC, product pages via ItemList JSON-LD ──
    {
        "name": "Hero Lectro",
        "platform": "herolectro",
        "index_url": "https://www.herolectro.com/bikes/",
    },
    # ── Giant India — custom HTML scraper (prices are JS-loaded but names/categories available) ──
    {
        "name": "Giant",
        "platform": "giant",
        "base_url": "https://www.giant-bicycles.com",
        "catalog_paths": [
            "/in/bikes/mountain-bikes/",
            "/in/bikes/road-bikes/",
            "/in/bikes/kids-bikes/",
        ],
    },
    # ── Btwin / Decathlon — Algolia search API ──
    {
        "name": "Btwin - Decathlon",
        "platform": "decathlon",
    },
    # ── Unavailable / no online store ────────────────────────────────────────
    {"name": "Montra",        "platform": "unavailable", "note": "montra.in returns empty redirect (57 bytes)"},
    {"name": "Hercules",      "platform": "unavailable", "note": "TI Cycles brand — website ti-cycles.com DNS failure"},
    {"name": "BSA",           "platform": "unavailable", "note": "TI Cycles brand — website DNS failure"},
    {"name": "Atlas Cycles",  "platform": "unavailable", "note": "atlascycles.co.in suspended"},
    {"name": "Ninety One",    "platform": "unavailable", "note": "ninetyone.in returned 530 error"},
    {"name": "Kross",         "platform": "unavailable", "note": "krossbikes.com DNS failure / site inaccessible"},
    {"name": "Mach City",     "platform": "unavailable", "note": "machcity.in DNS failure"},
    {"name": "Stryder",       "platform": "unavailable", "note": "stryder.in has no e-commerce product listings"},
]


def _normalize_url(url: str) -> str:
    u = (url or "").strip()
    if not u:
        return ""
    if not re.match(r"^https?://", u, re.I):
        u = "https://" + u
    return u.rstrip("/")


def _host(url: str) -> str:
    try:
        return (urlparse(url).netloc or "").lower().replace("www.", "")
    except Exception:
        return ""


def _variant_only_tail(txt: str) -> bool:
    t = re.sub(r"[\[\](){}]", " ", (txt or "").strip().lower())
    if not t:
        return False
    color_words = {
        "black", "white", "red", "blue", "green", "yellow", "orange", "pink",
        "purple", "silver", "grey", "gray", "gold", "matte", "gloss", "metallic",
        "teal", "brown", "beige", "maroon", "cyan",
    }
    size_pat = re.compile(r"^(?:xxs|xs|s|m|l|xl|xxl|xxxl|\d{2}(?:\.\d)?(?:\"|in|inch|cm|mm)?)$", re.I)
    words = [w for w in re.split(r"[^a-z0-9\"\.]+", t) if w]
    if not words:
        return False
    ok = 0
    for w in words:
        if w in color_words or size_pat.match(w) or w in {"size", "frame", "wheel"}:
            ok += 1
    return ok == len(words)


def _clean_model_name(name: str) -> str:
    n = re.sub(r"\s+", " ", (name or "")).strip(" -|,/\t\n\r")
    if not n:
        return ""

    n = n.replace("→", " ").replace("←", " ").strip()

    if "|" in n:
        n = n.split("|", 1)[0].strip()

    # Remove trailing parenthetical variant only: e.g. "Model X (Blue, M)"
    m = re.search(r"\(([^)]{1,40})\)\s*$", n)
    if m and _variant_only_tail(m.group(1)):
        n = n[:m.start()].strip(" -|,/")

    # Remove trailing dash suffix if it looks like color/size: e.g. "Model X - Metallic Green"
    if " - " in n:
        head, tail = n.rsplit(" - ", 1)
        if _variant_only_tail(tail):
            n = head.strip()

    # Trim explicit size suffixes: e.g. "Model X - Size M"
    n = re.sub(r"\s*[-,:]?\s*size\s*(?:xxs|xs|s|m|l|xl|xxl|xxxl|\d{2}(?:\.\d)?(?:\"|in|inch|cm|mm)?)\s*$", "", n, flags=re.I)
    # Remove trailing price snippets often embedded in cards.
    n = re.sub(r"\s+(?:from\s+)?(?:₹|\$|€|£|₱|INR|USD|EUR|GBP|PHP|AED)\s*[\d,.]+\s*$", "", n, flags=re.I)
    n = re.sub(r"\s{2,}", " ", n).strip(" -|,/")
    return n.strip()


def _shopify_probe(base_url: str) -> bool:
    base = _normalize_url(base_url)
    if not base:
        return False
    try:
        r = requests.get(f"{base}/products.json?limit=1", headers=HDR, timeout=12)
        if r.status_code != 200:
            return False
        j = r.json()
        return isinstance(j, dict) and "products" in j
    except Exception:
        return False


def _live_fx_rates_to_inr() -> dict:
    now = time.time()
    if FX_CACHE["rates"] and (now - FX_CACHE["ts"] < 6 * 3600):
        return FX_CACHE["rates"]

    rates = {**FX_FALLBACK}
    try:
        # Returns rates for 1 INR in target currencies; invert to get target->INR.
        r = requests.get("https://open.er-api.com/v6/latest/INR", timeout=8)
        data = r.json() if r.status_code == 200 else {}
        if data.get("result") == "success" and isinstance(data.get("rates"), dict):
            for code, per_inr in data["rates"].items():
                try:
                    per_inr = float(per_inr)
                    if per_inr > 0:
                        rates[code.upper()] = 1.0 / per_inr
                except Exception:
                    continue
            rates["INR"] = 1.0
    except Exception:
        pass

    FX_CACHE["ts"] = now
    FX_CACHE["rates"] = rates
    return rates


def _detect_currency_code(text: str, hint: str = "") -> str:
    t = (text or "").strip()
    h = (hint or "").strip().upper()
    if h in FX_FALLBACK or h == "INR":
        return h

    up = t.upper()
    if "₹" in t or "INR" in up or "RS" in up:
        return "INR"
    if "₱" in t or "PHP" in up:
        return "PHP"
    if "€" in t or "EUR" in up:
        return "EUR"
    if "£" in t or "GBP" in up:
        return "GBP"
    if "CHF" in up:
        return "CHF"
    if "A$" in up or "AUD" in up:
        return "AUD"
    if "C$" in up or "CAD" in up:
        return "CAD"
    if "S$" in up or "SGD" in up:
        return "SGD"
    if "AED" in up:
        return "AED"
    if "¥" in t or "JPY" in up:
        return "JPY"
    if "CNY" in up or "RMB" in up:
        return "CNY"
    if "$" in t or "USD" in up:
        return "USD"
    return "INR"


def _parse_amount(text: str) -> float:
    raw = str(text or "").strip()
    if not raw:
        return 0.0
    s = re.sub(r"[^\d,\.]", "", raw)
    if not s:
        return 0.0

    # Handle decimal separators robustly for mixed locales.
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        if re.search(r",\d{2}$", s):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")

    try:
        return float(s)
    except Exception:
        return 0.0


def _build_cfg_from_brand_website(brand: str, website: str) -> dict:
    brand = (brand or "").strip() or "Unknown Brand"
    website = _normalize_url(website)
    h = _host(website)
    b = brand.lower()

    # Brand-first overrides for dynamic / popup / region-gated sites.
    if "sportnetwork.in" in h and b in {"scott", "avanti", "bergamont"}:
        path_map = {
            "scott": "https://sportnetwork.in/brand-products-list/scott",
            "avanti": "https://sportnetwork.in/brand-products-list/Avanti",
            "bergamont": "https://sportnetwork.in/brand-products-list/bergamont",
        }
        return {
            "name": brand,
            "platform": "dynamic_models",
            "url": path_map[b],
            "currency_hint": "INR",
        }

    if b == "merida":
        return {
            "name": brand,
            "platform": "dynamic_models",
            "url": "https://www.merida-bikes.com/en/bikefinder/tag/bikes-83/root/bikes",
            "extra_paths": ["/en/bikefinder/archive/tag/bikes-83/root/archive"],
            "click_global": True,
        }

    if b in {"java", "felt", "specialized", "cervelo", "orbea"}:
        return {
            "name": brand,
            "platform": "dynamic_models",
            "url": website,
            "currency_hint": "PHP" if b == "java" else "",
        }

    # Exact known templates by host/name from existing BRANDS.
    for t in BRANDS:
        for key in ("base_url", "url", "index_url", "site_base"):
            v = t.get(key)
            if v and _host(v) == h:
                cfg = deepcopy(t)
                cfg["name"] = brand
                return cfg

    # Domain heuristics.
    if "trekbikes" in h:
        return {
            "name": brand,
            "platform": "occ",
            "base_url": "https://api.trekbikes.com",
            "site_base": "https://www.trekbikes.com",
            "store": "in",
            "lang": "en_IN",
            "root_category": "B100",
        }
    if "decathlon" in h:
        return {"name": brand, "platform": "decathlon"}
    if _shopify_probe(website):
        return {"name": brand, "platform": "shopify", "base_url": website}

    # Fallback: generic model-only scraper.
    return {"name": brand, "platform": "models_generic", "url": website}


def load_brands_from_excel(xlsx_path: str) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    headers = [str(ws.cell(1, c).value or "").strip().lower() for c in range(1, ws.max_column + 1)]
    try:
        brand_col = headers.index("brand") + 1
        web_col = headers.index("website") + 1
    except ValueError:
        raise ValueError("Input Excel must contain 'Brand' and 'Website' columns in row 1")

    items = []
    seen = set()
    for r in range(2, ws.max_row + 1):
        brand = str(ws.cell(r, brand_col).value or "").strip()
        web = str(ws.cell(r, web_col).value or "").strip()
        if not brand and not web:
            continue
        key = (brand.lower(), _normalize_url(web).lower())
        if key in seen:
            continue
        seen.add(key)
        items.append(_build_cfg_from_brand_website(brand, web))
    return items


def load_default_client_brands() -> list[dict]:
    items = []
    seen = set()
    for brand, website in DEFAULT_CLIENT_BRANDS:
        key = (brand.lower(), _normalize_url(website).lower())
        if key in seen:
            continue
        seen.add(key)
        items.append(_build_cfg_from_brand_website(brand, website))
    return items


def models_only_rows(rows: list[dict]) -> list[dict]:
    accessory_words = {
        "accessory", "accessories", "apparel", "helmet", "glove", "sock", "shoe",
        "jersey", "jacket", "bib", "bottle", "cage", "pump", "lock", "bell",
        "light", "mudguard", "fender", "tube", "tyre", "tire", "pedal", "saddle",
        "seatpost", "stem", "handlebar", "grip", "computer", "mount", "tool",
        "spare", "part", "brake", "cassette", "chain", "derailleur", "hub", "fork",
        "axle", "damper", "protector", "rack", "bag", "trainer", "nutrition",
        "component", "components", "frameset", "wheelset", "cockpit", "headset",
        "bottom bracket", "dropout", "chainring", "tape", "cap", "cover", "clamp",
    }
    bike_hint_words = {
        "bike", "bikes", "bicycle", "bicycles", "cycle", "cycles", "road", "mountain",
        "mtb", "gravel", "hybrid", "commuter", "city", "kids", "junior", "e-bike",
        "ebike", "cross country", "downhill", "trail", "triathlon",
    }
    generic_bucket_names = {
        "accessories", "apparel", "bikes", "bike accessories", "bicycle accessories",
        "bike services", "bike transport", "backpack", "bags", "helmets", "wheels",
        "find your bike", "reset filters", "see all bikes", "all bikes",
        "view details", "learn more", "shop now", "new bike registration", "2spin pre-owned",
        "facebook", "instagram", "youtube", "x", "twitter", "linkedin",
    }

    def _is_probable_bike_model(row: dict, name: str) -> bool:
        n = name.lower()
        cat = str(row.get("Category", "") or "").lower()
        tags = str(row.get("Tags", "") or "").lower()
        url = str(row.get("Product URL", "") or "").lower()

        if n in generic_bucket_names:
            return False
        if re.search(r"\b(series|registration|pre-owned|pre owned|find your bike|reset filters)\b", n):
            return False
        if "&" in n and not re.search(r"\d", n) and re.search(r"\b(performance|adventure|triathlon|gravel|all-terrain)\b", n):
            return False
        if len(n) <= 2:
            return False
        if len(n.split()) > 8 and not re.search(r"\d", n):
            return False
        if name.isupper() and len(name) <= 30 and not re.search(r"\d", name):
            return False

        blob = " ".join([n, cat, tags, url])
        has_bike_hint = any(w in blob for w in bike_hint_words)
        has_accessory_hint = any(w in blob for w in accessory_words)
        has_model_token = bool(re.search(r"[a-zA-Z].*\d|\d.*[a-zA-Z]", name))
        bike_segment_words = {
            "bike", "bikes", "cycle", "cycles", "mtb", "road", "gravel", "hybrid",
            "urban", "city", "kids", "junior", "triathlon", "bmx", "electric",
            "downhill", "cross country", "trail", "enduro",
        }
        category_has_bike_segment = any(w in cat for w in bike_segment_words)
        strong_non_bike_category = any(x in cat for x in {
            "spare", "parts", "apparel", "accessories", "helmet", "shoe", "component",
            "components", "brake", "cassette", "chain", "wheelset", "tool", "nutrition",
        }) and not category_has_bike_segment

        if strong_non_bike_category:
            return False

        # Real bike names often include technical terms (e.g., disc brakes, carbon fork).
        # Treat bike hints as stronger than accessory hints when the name looks model-like.
        if has_bike_hint:
            if has_accessory_hint and not category_has_bike_segment:
                return False
            return True

        if has_accessory_hint:
            return False

        # Fallback for sites that don't expose category hints.
        return has_model_token

    out = []
    seen = set()
    for row in rows:
        raw = str(row.get("Model Name", "") or "").strip()
        if not raw:
            continue
        if raw.startswith("["):
            # Preserve scrape error/unavailable notes for visibility in output sheets.
            out.append({
                "Model Name": raw,
                "Category": "",
                "Variant": "",
                "Current Price (Rs.)": "",
                "MRP / Strike Price (Rs.)": "",
                "Year": "",
                "Tags": "",
                "Product URL": "",
            })
            continue
        name = _clean_model_name(raw)
        if not name:
            continue
        if not _is_probable_bike_model(row, name):
            continue
        key = name.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append({
            "Model Name": name,
            "Category": row.get("Category", "") or "",
            "Variant": "",
            "Current Price (Rs.)": row.get("Current Price (Rs.)", "") or "",
            "MRP / Strike Price (Rs.)": row.get("MRP / Strike Price (Rs.)", "") or "",
            "Year": "",
            "Tags": "",
            "Product URL": row.get("Product URL", "") or "",
        })
    return out


def _output_value(row: dict, col_name: str):
    if col_name == "Model name":
        return row.get("Model Name", "")
    if col_name == "Segment":
        return row.get("Category", "")
    current = str(row.get("Current Price (Rs.)", "") or "")
    listed = str(row.get("MRP / Strike Price (Rs.)", "") or "")
    if col_name == "Listed price":
        return listed if listed else current
    if col_name == "Discounted price":
        if listed and current and current != listed:
            return current
        return ""
    return ""

# ══════════════════════════════════════════════════════════════════════════════
#  SHOPIFY SCRAPER
# ══════════════════════════════════════════════════════════════════════════════

def scrape_shopify(cfg: dict) -> list[dict]:
    base = cfg["base_url"].rstrip("/")
    rows, page = [], 1
    print(f"  [Shopify] {base}")
    while True:
        url = f"{base}/products.json?limit=250&page={page}"
        try:
            r = requests.get(url, headers=HDR, timeout=20)
        except Exception as e:
            print(f"    ERROR fetching page {page}: {e}")
            break
        if r.status_code != 200:
            print(f"    HTTP {r.status_code} on page {page} — stopping")
            break
        # Some Shopify stores redirect /products.json to a login / gated page
        ct = r.headers.get("Content-Type", "")
        if "json" not in ct:
            print(f"    Non-JSON response (Content-Type: {ct}) — site may require login or is not Shopify")
            break
        try:
            data = r.json()
        except Exception:
            print(f"    JSON parse error page {page} — trying with Accept header")
            # retry with explicit JSON accept header
            try:
                r2 = requests.get(url, headers={**HDR, "Accept": "application/json"}, timeout=20)
                data = r2.json()
            except Exception as e2:
                print(f"    Retry also failed: {e2}")
                break
        products = data.get("products", [])
        if not products:
            break
        for p in products:
            title   = p.get("title", "")
            p_type  = p.get("product_type", "")
            tags    = ", ".join(p.get("tags", []))
            handle  = p.get("handle", "")
            purl    = f"{base}/products/{handle}"
            variants = p.get("variants", [{}])
            published_at = p.get("published_at", "")
            year = published_at[:4] if published_at else ""
            for v in variants:
                price   = _to_rs(v.get("price", ""))
                compare = _to_rs(v.get("compare_at_price", ""))
                vtitle  = v.get("title", "")
                if vtitle == "Default Title":
                    vtitle = ""
                rows.append({
                    "Model Name":              title,
                    "Category":                p_type,
                    "Variant":                 vtitle,
                    "Current Price (Rs.)":     price,
                    "MRP / Strike Price (Rs.)": compare,
                    "Year":                    year,
                    "Tags":                    tags,
                    "Product URL":             purl,
                })
        print(f"    page {page}: {len(products)} products")
        page += 1
        time.sleep(0.5)
    return rows


# ══════════════════════════════════════════════════════════════════════════════
#  SAP COMMERCE CLOUD (OCC) — Trek
# ══════════════════════════════════════════════════════════════════════════════

def scrape_occ(cfg: dict) -> list[dict]:
    base      = cfg["base_url"].rstrip("/")
    store     = cfg.get("store", "in")
    lang      = cfg.get("lang", "en_IN")
    category  = cfg.get("root_category", "B100")  # Trek root bike category
    site_base = cfg.get("site_base", "https://www.trekbikes.com")
    rows, page, page_size = [], 0, 100
    occ_hdrs = {
        **HDR,
        "Accept":  "application/json",
        "Origin":  site_base,
        "Referer": site_base + "/",
    }
    print(f"  [OCC] {base}/occ/v2/{store}  category={category}")
    while True:
        url = (
            f"{base}/occ/v2/{store}/products/search?fields=FULL"
            f"&query=%3Arelevance%3Acategory%3A{category}"
            f"&lang={lang}&curr=INR&pageSize={page_size}&currentPage={page}"
        )
        data = None
        last_err = None
        for _ in range(3):
            try:
                r = requests.get(url, headers=occ_hdrs, timeout=20)
                r.raise_for_status()
                data = r.json()
                break
            except Exception as e:
                last_err = e
                time.sleep(1.0)
        if data is None:
            print(f"    ERROR: {last_err}")
            break
        prods = data.get("products", [])
        if not prods:
            break
        pagination = data.get("pagination", {})
        total      = pagination.get("totalResults", 0)
        total_pgs  = pagination.get("totalPages", 1)
        for p in prods:
            purl     = p.get("url", "")
            name     = p.get("name", "") or purl.rstrip("/").split("/")[-1].replace("-", " ").title()
            cat_name = (p.get("defaultCategory") or "").title()
            year     = str(p.get("marketingModelYear", "") or "")
            tag      = p.get("productCallout", "") or ""
            price_v  = (p.get("price") or {}).get("value", "")
            was_v    = (p.get("wasPrice") or {}).get("value", "")
            price    = _to_rs(str(price_v))
            mrp      = _to_rs(str(was_v)) if was_v and float(was_v or 0) > float(price_v or 0) else ""
            rows.append({
                "Model Name":              name,
                "Category":                cat_name,
                "Variant":                 "",
                "Current Price (Rs.)":     price,
                "MRP / Strike Price (Rs.)": mrp,
                "Year":                    year,
                "Tags":                    tag,
                "Product URL":             site_base + purl if purl.startswith("/") else purl,
            })
        print(f"    page {page+1}/{total_pgs}: {len(prods)} products  (total {total})")
        if page + 1 >= total_pgs:
            break
        page += 1
        time.sleep(0.4)
    if not rows and "trek" in str(cfg.get("name", "")).lower():
        print("    OCC returned no products; trying Trek sitemap fallback...")
        return scrape_trek_sitemap_models(cfg)
    return rows


def _trek_model_from_url(url: str) -> str:
    # Example: /in/en_IN/bikes/mountain-bikes/trail-mountain-bikes/fuel-ex/fuel-ex-8-gen-6/p/...
    path = urlparse(url).path.strip("/")
    parts = [p for p in path.split("/") if p]
    if not parts:
        return ""
    if "p" in parts:
        p_idx = parts.index("p")
        if p_idx > 0:
            slug = unquote(parts[p_idx - 1])
            return slug.replace("-", " ").strip().title()
    slug = parts[-1]
    slug = unquote(slug)
    return slug.replace("-", " ").strip().title()


def scrape_trek_sitemap_models(cfg: dict) -> list[dict]:
    site_base = cfg.get("site_base", "https://www.trekbikes.com")
    sitemap_index = site_base.rstrip("/") + "/in/en_IN/sitemap.xml"
    print(f"  [Trek fallback] {sitemap_index}")

    rows = []
    seen = set()

    try:
        idx_xml = requests.get(sitemap_index, headers=HDR, timeout=20).text
        sitemaps = re.findall(r"<loc>(.*?)</loc>", idx_xml)
        in_sitemaps = [u for u in sitemaps if "-en-IN-" in u]
        us_sitemaps = [u for u in sitemaps if "-en-US-" in u]
    except Exception as e:
        print(f"    Trek fallback error (index): {e}")
        return []

    for sm_url in in_sitemaps:
        try:
            sm_xml = requests.get(sm_url, headers=HDR, timeout=20).text
        except Exception:
            continue
        urls = re.findall(r"<loc>(.*?)</loc>", sm_xml)
        for u in urls:
            lu = u.lower()
            if "/bikes/" not in lu:
                continue
            if "/c/" in lu:
                continue
            model = _trek_model_from_url(u)
            if not model:
                continue
            key = model.lower()
            if key in seen:
                continue
            seen.add(key)
            rows.append({
                "Model Name": model,
                "Category": "",
                "Variant": "",
                "Current Price (Rs.)": "",
                "MRP / Strike Price (Rs.)": "",
                "Year": "",
                "Tags": "",
                "Product URL": u,
            })
        if len(rows) >= 250:
            break

    # India locale currently exposes very few model URLs in sitemap; backfill from en-US.
    if len(rows) < 30:
        for sm_url in us_sitemaps[:20]:
            try:
                sm_xml = requests.get(sm_url, headers=HDR, timeout=20).text
            except Exception:
                continue
            urls = re.findall(r"<loc>(.*?)</loc>", sm_xml)
            for u in urls:
                lu = u.lower()
                if "/bikes/" not in lu:
                    continue
                if "/c/" in lu:
                    continue
                model = _trek_model_from_url(u)
                if not model:
                    continue
                key = model.lower()
                if key in seen:
                    continue
                seen.add(key)
                rows.append({
                    "Model Name": model,
                    "Category": "",
                    "Variant": "",
                    "Current Price (Rs.)": "",
                    "MRP / Strike Price (Rs.)": "",
                    "Year": "",
                    "Tags": "",
                    "Product URL": u,
                })
            if len(rows) >= 250:
                break

    print(f"    Trek fallback models: {len(rows)}")
    return rows


# ══════════════════════════════════════════════════════════════════════════════
#  WOOCOMMERCE — HTML listing-page scraper
# ══════════════════════════════════════════════════════════════════════════════

_WC_PRODUCT_SEL  = "li.product, article.product, .product-item, .type-product"
_WC_NAME_SELS    = [
    ".woocommerce-loop-product__title", "h2.product-title",
    "h3.woocommerce-loop-product__title", "h3", "h2", ".product-name", ".name",
]
# Price selectors tried in order; if tag has a `content` attr, that takes priority
_WC_PRICE_SELS   = [
    ".price ins .woocommerce-Price-amount",
    "span.value",          # Hero Cycles / custom themes use content= attr
    ".price ins .amount",
    ".price .woocommerce-Price-amount",
    ".price .amount",
    ".woocommerce-Price-amount",
    ".amount",
    ".price",
]
_WC_MRP_SELS     = [".price del .woocommerce-Price-amount", ".price del .amount"]
_WC_LINK_SEL     = "a.woocommerce-loop-product__link, a.product-loop-image, a[href], a"
_WC_NEXT_SEL     = "a.next.page-numbers, .woocommerce-pagination a.next"


def _wc_price(card, selectors):
    """Try each selector; prefer 'content' attribute over text (Hero Cycles style)."""
    for sel in selectors:
        tag = card.select_one(sel)
        if not tag:
            continue
        if tag.get("content"):
            val = _to_rs(tag["content"])
            if val:
                return val
        txt = tag.get_text(separator=" ", strip=True)
        # skip pure labels like 'MRP', '₹', 'Price'
        nums = re.findall(r"[\d,]+", txt)
        if nums:
            return _to_rs(nums[-1].replace(",", ""))
    return ""


def _wc_name(card, selectors):
    for sel in selectors:
        tag = card.select_one(sel)
        if tag:
            name = tag.get_text(strip=True)
            if name and len(name) > 1:
                return name
    return ""


def scrape_woocommerce(cfg: dict) -> list[dict]:
    base       = cfg["base_url"].rstrip("/")
    paths      = cfg.get("catalog_paths", ["/shop/"])
    rows       = []
    seen_urls  = set()
    # Per-brand selector overrides
    name_sels  = cfg.get("name_sels",  _WC_NAME_SELS)
    price_sels = cfg.get("price_sels", _WC_PRICE_SELS)
    mrp_sels   = cfg.get("mrp_sels",   _WC_MRP_SELS)
    card_sel   = cfg.get("card_sel",   _WC_PRODUCT_SEL)
    link_sel   = cfg.get("link_sel",   _WC_LINK_SEL)
    print(f"  [WooCommerce] {base}")

    def _scrape_paged(start_url: str, category: str):
        url = start_url
        while url:
            try:
                r = requests.get(url, headers=HDR, timeout=20)
            except Exception as e:
                print(f"    ERR {url}: {e}")
                break
            if r.status_code not in (200, 301, 302):
                break
            soup = BeautifulSoup(r.text, "html.parser")
            cards = soup.select(card_sel)
            if not cards:
                cards = soup.find_all(class_=re.compile(r"\bproduct\b"))
            added = 0
            for card in cards:
                name  = _wc_name(card, name_sels)
                price = _wc_price(card, price_sels)
                mrp   = _wc_price(card, mrp_sels)
                link_tag = card.select_one(link_sel)
                href  = link_tag.get("href", "") if link_tag else ""
                purl  = href if href.startswith("http") else (base + href if href else "")

                if not name or purl in seen_urls:
                    continue
                seen_urls.add(purl)
                rows.append({
                    "Model Name":              name,
                    "Category":                category,
                    "Variant":                 "",
                    "Current Price (Rs.)":     price,
                    "MRP / Strike Price (Rs.)": mrp if mrp != price else "",
                    "Year":                    "",
                    "Tags":                    "",
                    "Product URL":             purl,
                })
                added += 1
            print(f"    {url}  → {added} products (found {len(cards)} cards)")
            nxt = soup.select_one(_WC_NEXT_SEL)
            url = nxt["href"] if nxt and nxt.get("href") else None
            time.sleep(0.5)

    for path in paths:
        category = path.strip("/").replace("-", " ").title()
        _scrape_paged(base + path, category)

    return rows


# ══════════════════════════════════════════════════════════════════════════════
#  SFCC / DEMANDWARE — Hero Lectro
# ══════════════════════════════════════════════════════════════════════════════

# Multiple selector sets to try (SFCC templates vary)
_SFCC_TILE_SELS = [
    ("article.product-tile, div.product-tile, li.product-grid-tile", "div.pdp-link a, a.link, h2 a", ".price-sales, .sales .value, del ~ .value, ins .value, .price-container .value"),
    (".c-product-tile, .product-item",  "h2 a, h3 a, .product-name a", ".price"),
    ("li.grid-tile, .grid-tile",        ".product-name a, a.thumb-link", ".product-pricing, .price-sales"),
]

def scrape_sfcc(cfg: dict) -> list[dict]:
    base  = cfg["base_url"].rstrip("/")
    paths = cfg.get("catalog_paths", ["/bikes/"])
    rows  = []
    seen  = set()
    print(f"  [SFCC] {base}")

    for path in paths:
        url = base + path
        try:
            r = requests.get(url, headers=HDR, timeout=20)
        except Exception as e:
            print(f"    ERR {url}: {e}")
            continue
        if r.status_code != 200:
            print(f"    HTTP {r.status_code} at {path}")
            continue
        soup = BeautifulSoup(r.text, "html.parser")
        category = path.strip("/").replace("-", " ").title()

        # Try each selector combo until we get products
        for (tile_sel, name_sel, price_sel) in _SFCC_TILE_SELS:
            tiles = soup.select(tile_sel)
            if tiles:
                print(f"    {url}  {len(tiles)} tiles via «{tile_sel}»")
                for tile in tiles:
                    name_tag  = tile.select_one(name_sel)
                    price_tag = tile.select_one(price_sel)
                    link_tag  = tile.select_one("a[href]")

                    name  = name_tag.get_text(strip=True)  if name_tag  else ""
                    price = _to_rs(price_tag.get_text(strip=True)) if price_tag else ""
                    href  = link_tag["href"] if link_tag else ""
                    purl  = href if href.startswith("http") else base + href

                    if not name or purl in seen:
                        continue
                    seen.add(purl)
                    rows.append({
                        "Model Name":              name,
                        "Category":                category,
                        "Variant":                 "",
                        "Current Price (Rs.)":     price,
                        "MRP / Strike Price (Rs.)": "",
                        "Year":                    "",
                        "Tags":                    "",
                        "Product URL":             purl,
                    })
                break  # found something, no need to try next selector combo
        else:
            # Fallback: try any anchor that looks like a product
            print(f"    {path}: no tiles matched, trying generic anchors")
            for a in soup.find_all("a", href=re.compile(r"/bikes/|/product|/p/")):
                name = a.get_text(strip=True)
                href = a["href"]
                purl = href if href.startswith("http") else base + href
                if len(name) > 4 and purl not in seen:
                    seen.add(purl)
                    rows.append({
                        "Model Name":              name,
                        "Category":                category,
                        "Variant":                 "",
                        "Current Price (Rs.)":     "",
                        "MRP / Strike Price (Rs.)": "",
                        "Year":                    "",
                        "Tags":                    "",
                        "Product URL":             purl,
                    })
        time.sleep(0.5)
    return rows


# ══════════════════════════════════════════════════════════════════════════════
#  GIANT INDIA — custom static HTML scraper
#  Names are in static HTML; prices are JS-loaded so left blank
# ══════════════════════════════════════════════════════════════════════════════

def scrape_giant(cfg: dict) -> list[dict]:
    base  = cfg["base_url"].rstrip("/")
    paths = cfg.get("catalog_paths", [])
    rows  = []
    seen  = set()
    print(f"  [Giant] {base}")
    for path in paths:
        url = base + path
        try:
            r = requests.get(url, headers=HDR, timeout=20)
        except Exception as e:
            print(f"    ERR {url}: {e}"); continue
        if r.status_code != 200:
            print(f"    HTTP {r.status_code} at {path}"); continue
        soup = BeautifulSoup(r.text, "html.parser")
        category = path.split("/")[-2].replace("-", " ").title()
        cards = soup.select(".product")
        added = 0
        for card in cards:
            # Giant: two <a> tags with same href; second one has the name text
            all_links = card.find_all("a", href=True)
            link = None
            name = ""
            for a in all_links:
                texts = [t.strip() for t in a.stripped_strings]
                noise = {"current price:", "view", "models", "time:"}
                name_parts = [t for t in texts if t.lower() not in noise
                              and not re.match(r"^\d+\s*(model|ms)s?$", t.lower())
                              and len(t) > 1]
                if name_parts:
                    name = name_parts[0]
                    link = a
                    break
            if not name or not link:
                continue
            href = link.get("href", "")
            if not href.startswith("http"):
                href = f"https://www.giant-bicycles.com{href}"
            if href in seen:
                continue
            seen.add(href)
            rows.append({
                "Model Name":              name,
                "Category":                category,
                "Variant":                 "",
                "Current Price (Rs.)":     "",   # JS-rendered, use website
                "MRP / Strike Price (Rs.)": "",
                "Year":                    "",
                "Tags":                    "",
                "Product URL":             href,
            })
            added += 1
        print(f"    {url}  → {added} products ({len(cards)} cards)")
        time.sleep(0.5)
    return rows


# ══════════════════════════════════════════════════════════════════════════════
#  HERO LECTRO — SFCC site: collect product URLs from ItemList JSON-LD,
#  then fetch each product page's Product JSON-LD for name + price
# ══════════════════════════════════════════════════════════════════════════════

def scrape_herolectro(cfg: dict) -> list[dict]:
    index_url = cfg.get("index_url", "https://www.herolectro.com/bikes/")
    rows = []
    print(f"  [Hero Lectro] {index_url}")
    try:
        r = requests.get(index_url, headers=HDR, timeout=20)
    except Exception as e:
        print(f"    ERR index: {e}"); return []
    if r.status_code != 200:
        print(f"    HTTP {r.status_code}"); return []

    soup = BeautifulSoup(r.text, "html.parser")
    product_urls = []
    for script in soup.find_all("script", type="application/ld+json"):
        try:
            d = json.loads(script.string)
            if d.get("@type") == "ItemList":
                for item in d.get("itemListElement", []):
                    u = item.get("url", "")
                    if u:
                        product_urls.append(u)
        except Exception:
            pass

    print(f"    Found {len(product_urls)} product URLs from ItemList")
    for purl in product_urls:
        try:
            pr = requests.get(purl, headers=HDR, timeout=20)
        except Exception as e:
            print(f"    ERR {purl}: {e}"); continue
        if pr.status_code != 200:
            continue
        psoup = BeautifulSoup(pr.text, "html.parser")
        name, price, mrp = "", "", ""
        for pscript in psoup.find_all("script", type="application/ld+json"):
            try:
                pd = json.loads(pscript.string)
                if pd.get("@type") == "Product":
                    name = pd.get("name", "")
                    offers = pd.get("offers", {})
                    if isinstance(offers, list): offers = offers[0]
                    price = _to_rs(str(offers.get("price", "")))
                    break
            except Exception:
                pass
        if not name:
            # fallback: page title
            t = psoup.find("title")
            name = t.get_text(strip=True).split("|")[0].strip() if t else ""
        rows.append({
            "Model Name":              name,
            "Category":                "E-Bike",
            "Variant":                 "",
            "Current Price (Rs.)":     price,
            "MRP / Strike Price (Rs.)": mrp,
            "Year":                    "",
            "Tags":                    "",
            "Product URL":             purl,
        })
        print(f"    {name}  ₹{price}")
        time.sleep(0.3)
    return rows


# ══════════════════════════════════════════════════════════════════════════════
#  DECATHLON / BTWIN — Algolia search API
# ══════════════════════════════════════════════════════════════════════════════

def scrape_decathlon(cfg: dict) -> list[dict]:
    """Fetch Btwin/Decathlon India cycles via Algolia search API."""
    ALGOLIA_APP_ID = "TH8IX1G260"
    ALGOLIA_API_KEY = "cd4f09aea8452f426737260804316800"
    INDEX_NAME = "dsi_pim_migration_category"
    BASE_URL = "https://www.decathlon.in/p/"

    # Categories that correspond to actual bicycles (not accessories)
    BIKE_CATS = [
        "Mountain Bikes", "Road Bikes", "Cycles for Women", "Cycles",
        "Cross Country Bikes", "Hybrid Cycles", "Btwin", "Kids Bikes",
        "Folding Bikes", "Electric Bikes", "BMX", "Cycle",
        "Gravel Bikes", "Trekking Bikes",
    ]
    cat_filter = " OR ".join(f'category_en:"{c}"' for c in BIKE_CATS)

    api_url = f"https://{ALGOLIA_APP_ID}-dsn.algolia.net/1/indexes/{INDEX_NAME}/query"
    headers = {
        "X-Algolia-Application-Id": ALGOLIA_APP_ID,
        "X-Algolia-API-Key": ALGOLIA_API_KEY,
        "Content-Type": "application/json",
    }

    rows = []
    seen = set()
    page = 0
    total_pages = 1

    print(f"  [Decathlon] Querying Algolia index for cycles...")
    while page < total_pages:
        try:
            r = requests.post(api_url, headers=headers, json={
                "query": "",
                "hitsPerPage": 100,
                "page": page,
                "filters": f"({cat_filter})",
            }, timeout=15)
            r.raise_for_status()
            data = r.json()
        except Exception as e:
            print(f"  [Decathlon] Algolia API error: {e}")
            break

        hits = data.get("hits", [])
        total_pages = data.get("nbPages", 1)
        for hit in hits:
            name = hit.get("name_en", "").strip()
            if not name or name.startswith("*"):
                # Strip leading * from internal-flagged names
                name = name.lstrip("* ").strip()
            if not name:
                continue
            key = name.lower()
            if key in seen:
                continue
            seen.add(key)

            price_raw = hit.get("price")
            mrp_raw = hit.get("price_mrp") or hit.get("price_nodiscount")
            price = float(price_raw) if price_raw else None
            mrp = float(mrp_raw) if mrp_raw else None

            slug = hit.get("link_en", "")
            product_url = (BASE_URL + slug) if slug else ""

            category = hit.get("category_en", "")
            rows.append({
                "Model Name": name,
                "Category": category,
                "Variant": "",
                "Current Price (Rs.)": price,
                "MRP / Strike Price (Rs.)": mrp,
                "Year": "",
                "Tags": "",
                "Product URL": product_url,
            })
        print(f"  [Decathlon] page {page+1}/{total_pages}: {len(hits)} hits")
        page += 1

    print(f"  [Decathlon] total cycles scraped: {len(rows)}")
    return rows


def _extract_jsonld_names(data, names: set[str]):
    if isinstance(data, dict):
        t = data.get("@type")
        t_str = " ".join(t) if isinstance(t, list) else str(t or "")
        t_str = t_str.lower()
        if "product" in t_str:
            n = _clean_model_name(str(data.get("name", "") or ""))
            if n:
                names.add(n)
        if "itemlist" in t_str:
            for it in data.get("itemListElement", []) or []:
                if isinstance(it, dict):
                    if isinstance(it.get("item"), dict):
                        n = _clean_model_name(str(it["item"].get("name", "") or ""))
                    else:
                        n = _clean_model_name(str(it.get("name", "") or ""))
                    if n:
                        names.add(n)
        for v in data.values():
            _extract_jsonld_names(v, names)
    elif isinstance(data, list):
        for it in data:
            _extract_jsonld_names(it, names)


def _extract_product_details_from_jsonld(data):
    """Return first Product-like tuple: (name, segment, current_price, listed_price)."""
    found = []

    def walk(node):
        if isinstance(node, dict):
            t = node.get("@type")
            t_str = " ".join(t) if isinstance(t, list) else str(t or "")
            t_str = t_str.lower()
            if "product" in t_str:
                name = _clean_model_name(str(node.get("name", "") or ""))
                segment = str(node.get("category", "") or "")
                cur = ""
                listed = ""

                offers = node.get("offers", {})
                if isinstance(offers, list) and offers:
                    offers = offers[0]
                if isinstance(offers, dict):
                    p_cur = str(offers.get("priceCurrency", "") or node.get("priceCurrency", ""))
                    cur = _to_rs(offers.get("price", ""), currency_hint=p_cur)
                    listed = _to_rs(offers.get("highPrice", ""), currency_hint=p_cur)
                    if not listed:
                        ps = offers.get("priceSpecification")
                        if isinstance(ps, dict):
                            listed = _to_rs(ps.get("price", ""), currency_hint=p_cur)
                        elif isinstance(ps, list):
                            for p in ps:
                                if isinstance(p, dict):
                                    listed = _to_rs(p.get("price", ""), currency_hint=p_cur)
                                    if listed:
                                        break
                if name:
                    found.append((name, segment, cur, listed))
            for v in node.values():
                walk(v)
        elif isinstance(node, list):
            for it in node:
                walk(it)

    walk(data)
    return found[0] if found else ("", "", "", "")


def scrape_models_generic(cfg: dict) -> list[dict]:
    base = _normalize_url(cfg.get("url", ""))
    print(f"  [Generic] {base}")
    if not base:
        return []

    paths = [
        "", "/bikes", "/bicycles", "/cycles", "/products", "/models",
        "/collections/all", "/collections/all-bikes", "/collections/bikes",
    ]
    candidate_urls = [base + p for p in paths]

    names = set()
    details = {}
    links = {}
    for u in candidate_urls:
        try:
            r = requests.get(u, headers=HDR, timeout=15)
            if r.status_code >= 400:
                continue
            soup = BeautifulSoup(r.text, "html.parser")

            # JSON-LD products / item lists.
            for s in soup.select("script[type='application/ld+json']"):
                raw = (s.string or s.get_text() or "").strip()
                if not raw:
                    continue
                try:
                    data = json.loads(raw)
                    _extract_jsonld_names(data, names)
                    d_name, d_seg, d_cur, d_listed = _extract_product_details_from_jsonld(data)
                    if d_name:
                        details[d_name] = {
                            "segment": d_seg,
                            "current": d_cur,
                            "listed": d_listed,
                        }
                except Exception:
                    pass

            # Product anchors/headings.
            sels = [
                "a[href*='product']", "a[href*='bike']", "a[href*='cycle']",
                ".product-title", ".product-name",
            ]
            for sel in sels:
                for el in soup.select(sel):
                    txt = _clean_model_name(el.get_text(" ", strip=True))
                    if not txt:
                        continue
                    if len(txt) < 3 or len(txt) > 120:
                        continue
                    # Ignore obvious navigation text.
                    if txt.lower() in {
                        "shop", "products", "bikes", "cycles", "read more", "learn more",
                        "view all", "buy now", "compare", "details",
                    }:
                        continue
                    names.add(txt)
                    href = el.get("href", "")
                    if href and txt not in links:
                        if href.startswith("/"):
                            links[txt] = base + href
                        elif href.startswith("http"):
                            links[txt] = href
        except Exception:
            continue

    rows = []
    for n in sorted(names):
        segment = details.get(n, {}).get("segment", "")
        current = details.get(n, {}).get("current", "")
        listed = details.get(n, {}).get("listed", "")
        purl = links.get(n, "")

        # Enrich missing prices by visiting detected product URL.
        if purl and (not current or not listed):
            try:
                pr = requests.get(purl, headers=HDR, timeout=12)
                if pr.status_code == 200:
                    psoup = BeautifulSoup(pr.text, "html.parser")
                    for s in psoup.select("script[type='application/ld+json']"):
                        raw = (s.string or s.get_text() or "").strip()
                        if not raw:
                            continue
                        try:
                            data = json.loads(raw)
                            d_name, d_seg, d_cur, d_listed = _extract_product_details_from_jsonld(data)
                            if d_seg and not segment:
                                segment = d_seg
                            if d_cur and not current:
                                current = d_cur
                            if d_listed and not listed:
                                listed = d_listed
                            if current and listed:
                                break
                        except Exception:
                            continue
            except Exception:
                pass

        rows.append({
            "Model Name": n,
            "Category": segment,
            "Variant": "",
            "Current Price (Rs.)": current,
            "MRP / Strike Price (Rs.)": listed,
            "Year": "",
            "Tags": "",
            "Product URL": purl,
        })
    return rows


def _parse_prices_from_text(text: str, currency_hint: str = "") -> tuple[str, str]:
    if not text:
        return "", ""
    # Capture values with explicit currency first; then fallback to large naked numbers.
    price_chunks = re.findall(
        r"(?:INR|USD|EUR|GBP|PHP|AED|CHF|A\$|C\$|S\$|₹|\$|€|£|₱)\s*\d[\d,\.]*",
        text,
        flags=re.I,
    )
    if not price_chunks:
        if re.search(r"\b(price|mrp|sale|offer|discount|from|now|was)\b", text, flags=re.I):
            price_chunks = re.findall(r"\b\d{4,}(?:[\.,]\d{2})?\b", text)
    vals = []
    for ch in price_chunks:
        v = _to_rs(ch, currency_hint=currency_hint)
        if v:
            vals.append(int(v))
    vals = [v for v in vals if v >= 1000]
    if not vals:
        return "", ""
    vals = sorted(set(vals), reverse=True)
    if len(vals) >= 2:
        listed, current = vals[0], vals[-1]
        if current < listed:
            return str(listed), str(current)
        return str(listed), ""
    return str(vals[0]), ""


def scrape_dynamic_models(cfg: dict) -> list[dict]:
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print("  [Dynamic] Playwright not installed, falling back to generic")
        return scrape_models_generic(cfg)

    base = _normalize_url(cfg.get("url", ""))
    if not base:
        return []

    currency_hint = cfg.get("currency_hint", "")
    urls = [base]
    for p in cfg.get("extra_paths", []) or []:
        urls.append(urljoin(base + "/", p))

    rows = []
    seen = set()
    print(f"  [Dynamic] {base}")

    def close_popups(page):
        sel_list = [
            "button[aria-label*='close' i]",
            "button[aria-label*='dismiss' i]",
            ".close, .modal-close, .mfp-close, .popup-close, .newsletter-close",
            "button:has-text('Close')",
            "button:has-text('No thanks')",
            "button:has-text('Accept')",
            "button:has-text('I agree')",
            "button:has-text('Continue')",
        ]
        for sel in sel_list:
            try:
                btn = page.query_selector(sel)
                if btn:
                    btn.click(timeout=1200)
                    page.wait_for_timeout(250)
            except Exception:
                pass

    try:
        pw_ctx = sync_playwright().start()
        browser = pw_ctx.chromium.launch(headless=True)
    except Exception as e:
        print(f"  [Dynamic] Playwright runtime unavailable ({e}); falling back to generic")
        try:
            pw_ctx.stop()  # type: ignore[name-defined]
        except Exception:
            pass
        return scrape_models_generic(cfg)

    try:
        context = browser.new_context(extra_http_headers={
            "Accept-Language": "en-IN,en;q=0.9",
            "User-Agent": HDR["User-Agent"],
        })
        page = context.new_page()

        for u in urls:
            try:
                page.goto(u, wait_until="domcontentloaded", timeout=45000)
                page.wait_for_timeout(1200)
            except Exception:
                continue

            close_popups(page)

            if cfg.get("click_global"):
                try:
                    for sel in [
                        "a:has-text('Global')",
                        "button:has-text('Global')",
                        "a:has-text('International')",
                    ]:
                        el = page.query_selector(sel)
                        if el:
                            el.click(timeout=1500)
                            page.wait_for_load_state("networkidle", timeout=10000)
                            break
                except Exception:
                    pass

            # Scroll to trigger lazy-loaded listings.
            for _ in range(8):
                try:
                    page.mouse.wheel(0, 2500)
                    page.wait_for_timeout(300)
                except Exception:
                    break

            # Click load-more style buttons a few times.
            for _ in range(10):
                clicked = False
                for sel in [
                    "button:has-text('Load more')",
                    "button:has-text('Show more')",
                    "a:has-text('Load more')",
                    "a:has-text('Show more')",
                ]:
                    try:
                        btn = page.query_selector(sel)
                        if btn:
                            btn.click(timeout=1200)
                            page.wait_for_timeout(500)
                            clicked = True
                            break
                    except Exception:
                        pass
                if not clicked:
                    break

            soup = BeautifulSoup(page.content(), "html.parser")

            # Parse JSON-LD Product details from rendered page.
            for s in soup.select("script[type='application/ld+json']"):
                raw = (s.string or s.get_text() or "").strip()
                if not raw:
                    continue
                try:
                    data = json.loads(raw)
                except Exception:
                    continue

                found = []
                def walk(node):
                    if isinstance(node, dict):
                        t = node.get("@type")
                        t_str = " ".join(t) if isinstance(t, list) else str(t or "")
                        if "product" in t_str.lower():
                            found.append(node)
                        for v in node.values():
                            walk(v)
                    elif isinstance(node, list):
                        for it in node:
                            walk(it)
                walk(data)

                for p in found:
                    name = _clean_model_name(str(p.get("name", "") or ""))
                    if not name:
                        continue
                    key = name.lower()
                    if key in seen:
                        continue
                    offers = p.get("offers", {})
                    if isinstance(offers, list) and offers:
                        offers = offers[0]
                    code = ""
                    current = ""
                    listed = ""
                    if isinstance(offers, dict):
                        code = str(offers.get("priceCurrency", "") or "")
                        current = _to_rs(offers.get("price", ""), currency_hint=code or currency_hint)
                        listed = _to_rs(offers.get("highPrice", ""), currency_hint=code or currency_hint)
                    seg = str(p.get("category", "") or "")
                    link = str(p.get("url", "") or "")
                    if link and link.startswith("/"):
                        link = urljoin(page.url, link)
                    seen.add(key)
                    rows.append({
                        "Model Name": name,
                        "Category": seg,
                        "Variant": "",
                        "Current Price (Rs.)": current,
                        "MRP / Strike Price (Rs.)": listed,
                        "Year": "",
                        "Tags": "",
                        "Product URL": link,
                    })

            # Parse product cards/anchors for sites without rich JSON-LD.
            for a in soup.select("a[href]"):
                href = a.get("href", "")
                if not href:
                    continue
                href_l = href.lower()
                if not any(k in href_l for k in ["/product", "/products", "bike", "bikes", "cycle"]):
                    continue

                name = _clean_model_name(a.get_text(" ", strip=True))
                if not name:
                    continue
                key = name.lower()
                if key in seen:
                    continue

                link = href if href.startswith("http") else urljoin(page.url, href)
                ctx = " ".join([
                    a.get_text(" ", strip=True),
                    a.parent.get_text(" ", strip=True) if a.parent else "",
                ])
                listed, discounted = _parse_prices_from_text(ctx, currency_hint=currency_hint)

                seen.add(key)
                rows.append({
                    "Model Name": name,
                    "Category": "",
                    "Variant": "",
                    "Current Price (Rs.)": discounted,
                    "MRP / Strike Price (Rs.)": listed,
                    "Year": "",
                    "Tags": "",
                    "Product URL": link,
                })

        browser.close()
        pw_ctx.stop()
    except Exception:
        try:
            browser.close()
            pw_ctx.stop()
        except Exception:
            pass
        return scrape_models_generic(cfg)

    if not rows:
        return scrape_models_generic(cfg)
    return rows


# ══════════════════════════════════════════════════════════════════════════════
#  PLAYWRIGHT (JS-rendered sites)
# ══════════════════════════════════════════════════════════════════════════════

def scrape_playwright(cfg: dict) -> list[dict]:
    try:
        from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
    except ImportError:
        print("  [Playwright] not installed — skipping", cfg["name"])
        return []

    rows = []
    seen = set()
    cfg_url     = cfg.get("url", "")
    product_sel = cfg.get("product_sel", "")
    name_sel    = cfg.get("name_sel", "")
    price_sel   = cfg.get("price_sel", "")
    link_sel    = cfg.get("link_sel", "a")
    next_sel    = cfg.get("next_page_selector", "")

    print(f"  [Playwright] {cfg_url}")
    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=True)
        page    = browser.new_page(extra_http_headers={
            "Accept-Language": "en-IN,en;q=0.9",
            "User-Agent": HDR["User-Agent"],
        })
        current_url = cfg_url
        page_no = 0
        while current_url and page_no < 20:
            page_no += 1
            try:
                page.goto(current_url, wait_until="networkidle", timeout=30000)
            except PWTimeout:
                print(f"    timeout on page {page_no}")
                break
            except Exception as e:
                print(f"    error: {e}")
                break

            content = page.content()
            soup    = BeautifulSoup(content, "html.parser")

            # Try configured selector first, then generic fallbacks
            for sel in [product_sel, "article, li.product, .product-item, [data-testid='product-card']"]:
                if not sel:
                    continue
                cards = soup.select(sel)
                if cards:
                    added = 0
                    for card in cards:
                        n_tag = card.select_one(name_sel) if name_sel else None
                        p_tag = card.select_one(price_sel) if price_sel else None
                        l_tag = card.select_one(link_sel) or card if link_sel else card

                        name  = n_tag.get_text(strip=True) if n_tag else card.get_text(" ", strip=True)[:60]
                        price = _to_rs(p_tag.get_text(strip=True)) if p_tag else ""
                        href  = l_tag.get("href", "") if hasattr(l_tag, "get") else ""
                        if not href.startswith("http"):
                            from urllib.parse import urljoin
                            href = urljoin(current_url, href)

                        if not name or href in seen:
                            continue
                        seen.add(href)
                        rows.append({
                            "Model Name":              name,
                            "Category":                "",
                            "Variant":                 "",
                            "Current Price (Rs.)":     price,
                            "MRP / Strike Price (Rs.)": "",
                            "Year":                    "",
                            "Tags":                    "",
                            "Product URL":             href,
                        })
                        added += 1
                    print(f"    page {page_no}: {added} products (selector: {sel[:30]})")
                    break

            # Try to go to next page
            if not next_sel:
                break
            nxt = soup.select_one(next_sel)
            if not nxt or not nxt.get("href"):
                # try clicking via playwright
                try:
                    btn = page.query_selector(next_sel)
                    if btn:
                        btn.click()
                        page.wait_for_load_state("networkidle", timeout=10000)
                        current_url = page.url
                        continue
                except Exception:
                    pass
                break
            from urllib.parse import urljoin
            current_url = urljoin(current_url, nxt["href"])

        browser.close()
    return rows


# ══════════════════════════════════════════════════════════════════════════════
#  HELPERS
# ══════════════════════════════════════════════════════════════════════════════
import sys
CONVERT_INR_FLAGS = "--inr" in sys.argv

def _to_rs(val: str, currency_hint: str = "") -> str:
    """Normalize arbitrary currency price to INR integer string."""
    if val is None or val == "":
        return ""
    amount = _parse_amount(str(val))
    if amount <= 0:
        return ""

    if not CONVERT_INR_FLAGS:
        # Return exact numeric amount without conversion
        return str(int(round(amount)))

    code = _detect_currency_code(str(val), hint=currency_hint)
    rates = _live_fx_rates_to_inr()
    mult = rates.get(code, 1.0)
    try:
        inr = int(round(amount * float(mult)))
        return str(inr) if inr > 0 else ""
    except Exception:
        return str(int(round(amount)))


def _note_row(note: str) -> list[dict]:
    return [{"Model Name": f"[{note}]",
             "Category": "", "Variant": "",
             "Current Price (Rs.)": "", "MRP / Strike Price (Rs.)": "",
             "Year": "", "Tags": "", "Product URL": ""}]


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL EXPORT
# ══════════════════════════════════════════════════════════════════════════════

HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
HDR_FONT  = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL  = PatternFill("solid", fgColor="D6E4F0")
WARN_FILL = PatternFill("solid", fgColor="FCE4D6")

def export_to_excel(all_results: dict, out_path: str):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    summary_ws = wb.create_sheet("Summary", 0)
    summary_ws.append(["Brand", "Products Found", "Status"])
    for cell in summary_ws[1]:
        cell.font  = HDR_FONT
        cell.fill  = HDR_FILL
        cell.alignment = Alignment(horizontal="center")

    for brand_name, rows in all_results.items():
        ok     = rows and not rows[0]["Model Name"].startswith("[")
        status = "✓ Scraped" if ok else "✗ Unavailable"
        count  = len(rows) if ok else 0
        summary_ws.append([brand_name, count, status])
        r = summary_ws.max_row
        if not ok:
            for c in summary_ws[r]:
                c.fill = WARN_FILL

        # product sheet (strip chars invalid in Excel sheet names; max 31 chars)
        _invalid = r'[\[\]\*\/\?\:\\]'
        sheet_name = re.sub(_invalid, '-', brand_name)[:31]
        ws = wb.create_sheet(sheet_name)

        # header row
        for col_i, col_name in enumerate(COLS, 1):
            cell = ws.cell(row=1, column=col_i, value=col_name)
            cell.font      = HDR_FONT
            cell.fill      = HDR_FILL
            cell.alignment = Alignment(horizontal="center")

        for row_i, row in enumerate(rows, 2):
            fill = ALT_FILL if row_i % 2 == 0 else None
            for col_i, col_name in enumerate(COLS, 1):
                cell = ws.cell(row=row_i, column=col_i, value=_output_value(row, col_name))
                if fill:
                    cell.fill = fill

        # auto-width
        for col_i, col_name in enumerate(COLS, 1):
            max_len = len(col_name)
            for row in ws.iter_rows(min_row=2, min_col=col_i, max_col=col_i):
                val = str(row[0].value or "")
                max_len = max(max_len, len(val))
            ws.column_dimensions[get_column_letter(col_i)].width = min(max_len + 4, 60)

        ws.freeze_panes = "A2"
        print(f"  Sheet «{sheet_name}» — {len(rows)} rows")

    # fix summary column widths
    for col in summary_ws.columns:
        w = max(len(str(c.value or "")) for c in col)
        summary_ws.column_dimensions[get_column_letter(col[0].column)].width = w + 4

    wb.save(out_path)
    print(f"\n✅  Saved → {out_path}")


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════

def scrape_brand(cfg: dict) -> list[dict]:
    p = cfg["platform"]
    try:
        if p == "shopify":
            return scrape_shopify(cfg)
        if p == "occ":
            return scrape_occ(cfg)
        if p == "woocommerce":
            return scrape_woocommerce(cfg)
        if p == "sfcc":
            return scrape_sfcc(cfg)
        if p == "giant":
            return scrape_giant(cfg)
        if p == "herolectro":
            return scrape_herolectro(cfg)
        if p == "decathlon":
            return scrape_decathlon(cfg)
        if p == "dynamic_models":
            return scrape_dynamic_models(cfg)
        if p == "models_generic":
            return scrape_models_generic(cfg)
        if p == "playwright":
            return scrape_playwright(cfg)
        if p == "unavailable":
            return _note_row(cfg.get("note", "Website unavailable"))
    except Exception as e:
        print(f"  !! Exception: {e}")
        return _note_row(f"Scrape error: {e}")
    return []


def main():
    print("=" * 60)
    print(" Indian Bicycle Brands — Multi-sheet Price Scraper")
    print("=" * 60)

    # Usage:
    #   python scrape_all_brands.py
    #   python scrape_all_brands.py "bicycle competition analysis.xlsx"
    #   python scrape_all_brands.py --models-only
    args = sys.argv[1:]
    
    out_file = "all_bike_models.xlsx"
    if "--out" in args:
        idx = args.index("--out")
        out_file = args[idx + 1]
        args.pop(idx)
        args.pop(idx)
        
    models_only = True
    input_xlsx = ""
    filters = []

    for a in args:
        if a == "--models-only":
            models_only = True
        elif a == "--inr":
            pass
        elif a.lower().endswith(".xlsx") and Path(a).exists():
            input_xlsx = a
        else:
            filters.append(a.lower())

    brand_list = load_default_client_brands()
    if input_xlsx:
        print(f"\nInput sheet mode: {input_xlsx}")
        brand_list = load_brands_from_excel(input_xlsx)
        models_only = True
        print(f"Loaded {len(brand_list)} brands from input sheet")
    else:
        print(f"\nDefault client brand mode: {len(brand_list)} brands")

    all_results: dict[str, list[dict]] = {}

    for cfg in brand_list:
        name = cfg["name"]
        if filters and not any(f in name.lower() for f in filters):
            continue
        print(f"\n── {name} ──")
        rows = scrape_brand(cfg)
        if models_only:
            rows = models_only_rows(rows)
            if not rows:
                rows = _note_row("No model names detected")
        all_results[name] = rows
        print(f"  → {len(rows)} rows collected")

    out_name = out_file
    out = Path(__file__).parent / out_name
    export_to_excel(all_results, str(out))


if __name__ == "__main__":
    main()
