"""
Competition brand scraper using explicit method matrix.

Output columns per sheet:
- Model name
- Segment
- Listed price
- Discounted price

Usage:
  python scrape_competition_matrix.py
  python scrape_competition_matrix.py Trek Giant
"""

import json
import gzip
import re
import sys
import time
import sqlite3
from datetime import datetime
from urllib.parse import urljoin, urlparse, unquote
from concurrent.futures import ThreadPoolExecutor, as_completed

import openpyxl
import random
from curl_cffi import requests
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HDR = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-IN,en;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

COLS = ["Model name", "Segment", "Listed price", "Discounted price"]
OUT_XLSX = "competition_bike_models.xlsx"

STRICT_PHASE2_BRANDS = {
    "scott", "avanti", "bergamont", "merida", "orbea", "cube", "java", "jamis", "basso"
}

# Method matrix from user-provided logic.
BRANDS = [
    {"name": "Trek", "url": "https://bumsonthesaddle.com/collections/trek-bikes", "method": "shopify"},

    {"name": "Felt", "url": "https://feltbicycles.com/collections/all-bikes", "method": "shopify"},
    {"name": "Fuji", "url": "https://www.fujibikes.com/collections/all-bikes", "method": "shopify"},
    {"name": "Surly", "url": "https://surlybikes.com/", "method": "shopify"},
    {"name": "Kona", "url": "https://konaworld.com/", "method": "shopify"},
    {"name": "Marin", "url": "https://www.marinbikes.com/", "method": "shopify"},
    {"name": "Java", "url": "https://javabikesph.com/", "method": "shopify"},
    {"name": "Jamis", "url": "https://www.jamisbikes.com/int/", "method": "shopify"},
    {"name": "Basso", "url": "https://bassobikes.com/en", "method": "shopify"},

    {"name": "Canyon", "url": "https://www.canyon.com/en-in/", "method": "canyon_json"},

    {"name": "Scott", "url": "https://sportnetwork.in/", "method": "sportnetwork_brand"},
    {"name": "Avanti", "url": "https://sportnetwork.in/", "method": "sportnetwork_brand"},
    {"name": "Bergamont", "url": "https://sportnetwork.in/", "method": "sportnetwork_brand"},

    {"name": "BMC", "url": "https://bmc-switzerland.com/", "method": "html_jsonld"},
    {"name": "Cervelo", "url": "https://www.cervelo.com/en-us", "method": "html_jsonld"},
    {"name": "Pinarello", "url": "https://pinarello.com/usa/en", "method": "html_jsonld"},
    {"name": "Look", "url": "https://www.lookcycle.com/qe-en/", "method": "html_jsonld"},
    {"name": "Ridley", "url": "https://www.ridley-bikes.com/en_IN", "method": "html_jsonld"},
    {"name": "Factor", "url": "https://factorbikes.com/", "method": "html_jsonld"},
    {"name": "Orbea", "url": "https://www.orbea.com/in-en/", "method": "html_jsonld"},
    {"name": "Merida", "url": "https://www.merida-bikes.com/en", "method": "html_jsonld"},
    {"name": "Specialized", "url": "https://www.specialized.com/us/en", "method": "html_jsonld"},
    {"name": "Cannondale", "url": "https://www.cannondale.com/en", "method": "html_jsonld"},
    {"name": "Polygon", "url": "https://www.polygonbikes.com/", "method": "html_jsonld"},
    {"name": "Giant", "url": "https://giantindia.com/", "method": "html_jsonld"},
    {"name": "Cube", "url": "https://www.cube.eu/", "method": "html_jsonld"},
]

SHOPIFY_CURRENCY_HINT = {
    "felt": "USD",
    "fuji": "USD",
    "surly": "USD",
    "kona": "USD",
    "marin": "USD",
    "java": "PHP",
    "jamis": "USD",
    "basso": "EUR",
}

FX_FALLBACK = {
    "INR": 1.0,
    "USD": 83.0,
    "EUR": 90.0,
    "GBP": 105.0,
    "CHF": 94.0,
    "AED": 22.6,
    "PHP": 1.45,
}
FX_CACHE = {"ts": 0.0, "rates": {"INR": 1.0}}


def _safe_get(url: str, timeout: int = 25):
    return requests.get(url, headers=HDR, timeout=timeout)


def _host(url: str) -> str:
    return (urlparse(url).netloc or "").lower().replace("www.", "")


def _normalize_model(name: str) -> str:
    n = re.sub(r"\s+", " ", (name or "")).strip(" -|,/\t\n\r")
    if not n:
        return ""
    n = unquote(n)
    n = n.replace("→", " ").replace("←", " ")
    n = re.sub(r"(?i)^see the bike\s*[-–]?\s*", "", n)
    n = re.sub(r"\s*\|\s*.*$", "", n)
    n = re.sub(r"\s*[-,:]?\s*size\s*(?:xxs|xs|s|m|l|xl|xxl|xxxl|\d{2}(?:\.\d)?(?:\"|in|inch|cm|mm)?)\s*$", "", n, flags=re.I)
    n = re.sub(r"\s+\((?:black|white|red|blue|green|yellow|silver|grey|gray|gold|matte|metallic|xs|s|m|l|xl|xxl|\d+\")\)\s*$", "", n, flags=re.I)
    n = re.sub(r"\s+(?:from\s+)?(?:₹|\$|€|£|₱|INR|USD|EUR|GBP|PHP|AED)\s*[\d,.]+\s*$", "", n, flags=re.I)
    n = re.sub(r"\s{2,}", " ", n).strip(" -|,/")
    return n


def _looks_like_cycle_model(name: str) -> bool:
    t = (name or "").strip().lower()
    if not t:
        return False
    bad = [
        "helmet", "jersey", "glove", "socks", "shoe", "pedal", "saddle",
        "chain", "cassette", "brake", "rotor", "tire", "tyre", "tube", "pump",
        "bottle", "bidon", "tool", "grip", "bar tape", "computer", "light",
        "mudguard", "fender", "rack", "basket", "stand", "trainer", "nutrition",
        "cap", "t-shirt", "short", "jacket", "hoodie", "pants", "apparel",
        "kit", "gear", "derailleur", "shifter", "suspension", "fork"
    ]
    if any(b in t for b in bad):
        return False
    if len(t) < 3:
        return False
    return True


def _parse_price_num(txt):
    if txt is None:
        return None
    if isinstance(txt, dict):
        txt = txt.get("price") or txt.get("lowPrice") or txt.get("value")
        if txt is None: return None
        
    s = str(txt)
    # Exclude HTML/JSON blobs that leaked in
    if len(s) > 200: 
        return None
        
    m = re.search(r"(?:₹|\$|€|£|₱|INR|USD|EUR|GBP|PHP|AED)?\s*([\d][\d,]*(?:\.\d{1,2})?)", s)
    if not m:
        return None
    try:
        val = float(m.group(1).replace(",", ""))
        
        # Very small values are likely weights (e.g. 10.32kg -> 10.32) or model numbers
        if val < 50 and len(s) > 15:
            return None
            
        # Huge numbers are likely IDs or barcodes (e.g. GTIN/EAN)
        if val > 10000000:
            return None

        return val
    except Exception:
        return None

def fetch_price_from_search(brand, model_name):
    query = f"{brand} {model_name} cycle price india INR"
    url = f"https://html.duckduckgo.com/html/?q={query.replace(' ', '+')}"
    try:
        # curl_cffi impersonating chrome110 bypasses blocks brilliantly
        r = requests.get(url, impersonate="chrome110", timeout=15)
        if r.status_code == 200:
            soup = BeautifulSoup(r.text, "html.parser")
            for snip in soup.select(".result__snippet"):
                text = snip.get_text().lower()
                matches = re.findall(r"(?:rs\.?|₹|inr)\s*[\d,]+", text)
                if matches:
                    p = _parse_price_num(matches[0])
                    if p and p > 8000: return p
                m2 = re.search(r"price.+?starts? at\s*([\d,]+)", text)
                if m2:
                    p = _parse_price_num(m2.group(1))
                    if p and p > 8000: return p
    except Exception as e:
        pass
    return None


def _detect_currency(text: str, host: str = "") -> str:
    t = (text or "").upper()
    if "₹" in t or " INR" in t:
        return "INR"
    if "€" in t or " EUR" in t:
        return "EUR"
    if "£" in t or " GBP" in t:
        return "GBP"
    if "₱" in t or " PHP" in t:
        return "PHP"
    if "$" in t or " USD" in t:
        return "USD"

    h = (host or "").lower()
    if h.endswith(".in") or "/en-in" in h:
        return "INR"
    if h.endswith(".ph") or "ph" in h:
        return "PHP"
    return "USD"


def _get_fx_rates():
    now = time.time()
    if now - FX_CACHE["ts"] < 3600 and FX_CACHE["rates"]:
        return FX_CACHE["rates"]

    urls = [
        "https://open.er-api.com/v6/latest/INR",
        "https://api.exchangerate.host/latest?base=INR",
    ]
    for u in urls:
        try:
            j = requests.get(u, headers=HDR, timeout=15).json()
            rates = j.get("rates", {})
            if rates:
                out = {"INR": 1.0}
                for c in ["USD", "EUR", "GBP", "CHF", "AED", "PHP"]:
                    rv = rates.get(c)
                    if rv:
                        out[c] = 1.0 / float(rv)
                if len(out) > 1:
                    FX_CACHE["ts"] = now
                    FX_CACHE["rates"] = out
                    return out
        except Exception:
            pass

    FX_CACHE["ts"] = now
    FX_CACHE["rates"] = dict(FX_FALLBACK)
    return FX_CACHE["rates"]


CONVERT_INR = "--inr" in sys.argv

def _to_inr(price, currency: str):
    if price is None:
        return None
    cur = (currency or "INR").upper()
    if cur == "INR":
        return round(float(price), 2)
    
    if not CONVERT_INR:
        # User wants original currency numeric amount
        return round(float(price), 2)

    rates = _get_fx_rates()
    rate = rates.get(cur)
    if not rate:
        return round(float(price), 2)
    return round(float(price) * float(rate), 2)


def _row(model: str, segment: str = "", listed=None, discounted=None):
    return {
        "Model name": model,
        "Segment": segment or "",
        "Listed price": listed if listed is not None else "",
        "Discounted price": discounted if discounted is not None else "",
    }


def _model_from_url(url: str):
    path = urlparse(url).path.strip("/")
    if not path:
        return ""
    parts = [p for p in path.split("/") if p]
    if not parts:
        return ""
    if "p" in parts:
        i = parts.index("p")
        if i > 0:
            return _normalize_model(parts[i - 1].replace("-", " "))
    last = parts[-1]
    if last.endswith(".html"):
        if len(parts) > 1:
            return _normalize_model(parts[-2].replace("-", " "))
        return _normalize_model(last.replace(".html", "").replace("-", " "))
    if last.isdigit() and len(parts) > 1:
        return _normalize_model(parts[-2].replace("-", " "))
    return _normalize_model(last.replace("-", " "))


def _parse_json_payload(text: str):
    """Parse JSON even when endpoint prepends whitespace/noise."""
    if not text:
        return None
    start = text.find("{")
    end = text.rfind("}")
    if start < 0 or end <= start:
        return None
    try:
        return json.loads(text[start:end + 1])
    except Exception:
        return None


def scrape_sitemap_models(base_url: str, brand_name: str = ""):
    rows = []
    brand_l = (brand_name or "").lower()
    try:
        root = base_url.rstrip("/") + "/sitemap.xml"
        idx = _safe_get(root).text
    except Exception:
        return rows

    smaps = re.findall(r"<loc>(.*?)</loc>", idx)
    if not smaps:
        smaps = [root]

    allow_kw = ["bike", "bikes", "cycle", "road", "mountain", "gravel", "mtb", "hybrid"]
    deny_kw = ["helmet", "jersey", "shoe", "glove", "socks", "accessories", "apparel", "parts"]
    seen = set()

    for sm in smaps[:80]:
        try:
            xml = _safe_get(sm).text
        except Exception:
            continue
        urls = re.findall(r"<loc>(.*?)</loc>", xml)
        for u in urls:
            lu = u.lower()
            is_sportnetwork_brand = brand_l in {"scott", "avanti", "bergamont"} and "sportnetwork.in" in lu
            if brand_l in {"scott", "avanti", "bergamont"} and "sportnetwork.in" in lu:
                if "/product/" not in lu and "/products/" not in lu:
                    continue
            if not is_sportnetwork_brand and not any(k in lu for k in allow_kw):
                continue
            if any(k in lu for k in deny_kw):
                continue
            model = _model_from_url(u)
            if not _looks_like_cycle_model(model):
                continue
            key = model.lower()
            if key in seen:
                continue
            seen.add(key)
            rows.append(_row(model, "", "", ""))
            if len(rows) >= 1200:
                return rows
    return rows


def _sportnetwork_brand_seed_urls(brand_name: str):
    brand_l = (brand_name or "").strip().lower()
    seeds = [
        _sportnetwork_brand_url(brand_name),
        "https://sportnetwork.in/products-list/bike/bikes/mountain",
        "https://sportnetwork.in/products-list/bike/bikes/road",
        "https://sportnetwork.in/products-list/bike/bikes/gravel-cyclocross",
        "https://sportnetwork.in/products-list/bike/bikes/city-urban-hybrid",
        "https://sportnetwork.in/products-list/bike/bikes/e-bikes",
        "https://sportnetwork.in/products-list/bike/bikes/kids",
        "https://sportnetwork.in/products-list/bike/bikes/framesets",
    ]
    if brand_l == "avanti":
        # Avanti has mixed-case routes and sometimes links from nested bike pages.
        seeds.extend([
            "https://sportnetwork.in/brand-products-list/Avanti",
            "https://sportnetwork.in/brand-products-list/avanti",
            "https://sportnetwork.in/products-list/bike/demo/demo-bikes",
            "https://sportnetwork.in/products-list/bike/2ndspin-preowned/sportnetwork-certified",
        ])
    return list(dict.fromkeys(seeds))


def _extract_brand_product_slugs(text: str, brand_l: str):
    out = set()
    if not text:
        return out
    pat = re.compile(r"/brand-product-page/([a-z0-9-]+)/([^\"'?#\s<>]+)", flags=re.I)
    for b, slug in pat.findall(text):
        if (b or "").lower() != brand_l:
            continue
        slug = (slug or "").strip("/")
        if slug:
            out.add(slug)
    return out


def scrape_cube_sitemap(cfg):
    rows = []
    seen = set()
    try:
        idx = _safe_get("https://www.cube.eu/sitemap.xml").text
        locs = re.findall(r"<loc>(.*?)</loc>", idx)
    except Exception:
        return rows
    if not locs:
        return rows

    for sm in locs[:4]:
        try:
            raw = requests.get(sm, headers=HDR, timeout=35).content
            xml = gzip.decompress(raw).decode("utf-8", "ignore") if sm.endswith(".gz") else raw.decode("utf-8", "ignore")
        except Exception:
            continue

        urls = re.findall(r"<loc>(.*?)</loc>", xml)
        for u in urls:
            lu = u.lower()
            if any(x in lu for x in ["/faq/", "/dealer", "/service", "/jobs", "/news", "/blog", "/story"]):
                continue
            if not any(k in lu for k in ["/bike", "/bikes", "mountain", "road", "gravel", "hybrid", "emtb"]):
                continue
            model = _model_from_url(u)
            if not _looks_like_cycle_model(model):
                continue
            key = model.lower()
            if key in seen:
                continue
            seen.add(key)

            seg = ""
            if "mountain" in lu or "mtb" in lu:
                seg = "Mountain"
            elif "road" in lu:
                seg = "Road"
            elif "gravel" in lu:
                seg = "Gravel"
            elif "hybrid" in lu or "trekking" in lu or "tour" in lu:
                seg = "Hybrid/Trekking"
            rows.append(_row(model, seg, "", ""))

            if len(rows) >= 450:
                return rows
    return rows


def scrape_merida_bikefinder(cfg):
    rows = []
    seen = set()
    try:
        html = _safe_get("https://www.merida-bikes.com/en/bikefinder").text
    except Exception:
        return rows

    links = re.findall(r"/en/bike/\d+(?:-\d+)?/[^\"'\s<]+", html)
    if not links:
        return rows

    for p in links:
        model = _model_from_url(p)
        if not _looks_like_cycle_model(model):
            continue
        key = model.lower()
        if key in seen:
            continue
        seen.add(key)
        rows.append(_row(model, "", "", ""))
    return rows


def scrape_java_sitemap(cfg):
    rows = []
    seen = set()
    try:
        xml = _safe_get("https://javabikesph.com/sitemap.xml").text
    except Exception:
        return rows

    urls = re.findall(r"<loc>(.*?)</loc>", xml)
    for u in urls:
        lu = u.lower()
        if "/bikes/" not in lu:
            continue
        model = _model_from_url(u)
        if not _looks_like_cycle_model(model):
            continue
        key = model.lower()
        if key in seen:
            continue
        seen.add(key)
        rows.append(_row(model, "", "", ""))
    return rows


def scrape_jamis_sitemap(cfg):
    rows = []
    seen = set()
    seeds = [
        "https://www.jamisbikes.com/int/page-sitemap.xml",
        "https://www.jamisbikes.com/page-sitemap.xml",
    ]
    for sm in seeds:
        try:
            xml = _safe_get(sm).text
        except Exception:
            continue
        urls = re.findall(r"<loc>(.*?)</loc>", xml)
        for u in urls:
            lu = u.lower()
            if not any(k in lu for k in ["bike", "bikes", "road", "mountain", "gravel", "hybrid", "urban"]):
                continue
            if any(k in lu for k in ["faq", "event", "dealer", "contact", "about", "warranty"]):
                continue
            model = _model_from_url(u)
            if not _looks_like_cycle_model(model):
                continue
            key = model.lower()
            if key in seen:
                continue
            seen.add(key)
            rows.append(_row(model, "", "", ""))
    return rows


def scrape_orbea_playwright(cfg):
    rows = []
    seen = set()
    try:
        from playwright.sync_api import sync_playwright
    except Exception:
        return rows

    urls = [
        "https://www.orbea.com/us-en/bicycles/",
        "https://www.orbea.com/in-en/",
    ]
    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=True)
        page = browser.new_page(extra_http_headers={"Accept-Language": "en-US,en;q=0.9"})
        for u in urls:
            try:
                page.goto(u, wait_until="domcontentloaded", timeout=60000)
                page.wait_for_timeout(4500)
                for _ in range(6):
                    page.mouse.wheel(0, 5000)
                    page.wait_for_timeout(900)
                hrefs = page.eval_on_selector_all("a[href]", "els => els.map(e => e.getAttribute('href'))")
            except Exception:
                continue

            for h in hrefs or []:
                if not h:
                    continue
                full = urljoin(u, h)
                lu = full.lower()
                if not any(k in lu for k in ["/bicycles/", "/bike/"]):
                    continue
                if any(k in lu for k in ["/accessories/", "/equipment/", "/apparel/"]):
                    continue
                model = _model_from_url(full)
                if not _looks_like_cycle_model(model):
                    continue
                key = model.lower()
                if key in seen:
                    continue
                seen.add(key)
                rows.append(_row(model, "", "", ""))
        browser.close()
    return rows


def scrape_sportnetwork_api_brand(brand_name: str):
    rows = []
    seen = set()
    brand_l = brand_name.lower()
    combos = [
        ("bikes", "mountain"),
        ("bikes", "road"),
        ("bikes", "gravel-cyclocross"),
        ("bikes", "kids"),
        ("bikes", "framesets"),
        ("bikes", "city-urban-hybrid"),
        ("bikes", "e-bikes"),
        ("bikes", "limited-edition"),
        ("demo", "demo-bikes"),
        ("2ndspin-preowned", "sportnetwork-certified"),
        ("2ndspin-preowned", "dealer-pre-owned"),
    ]

    for subcat, sub in combos:
        for page in range(1, 26):
            api = (
                "https://dc.sportnetwork.in/products/v1/products_v1/index"
                f"?where[fl_item_categories.category_slug]=bike"
                f"&where[fl_item_categories.sub_category_slug]={subcat}"
                f"&where[fl_item_categories.sub_sub_category_slug]={sub}"
                f"&page_no={page}"
            )
            try:
                r = requests.get(api, headers={
                    **HDR,
                    "Referer": f"https://sportnetwork.in/products-list/bike/{subcat}/{sub}",
                    "Origin": "https://sportnetwork.in",
                    "X-Requested-With": "XMLHttpRequest",
                    "Accept": "application/json,text/plain,*/*",
                }, timeout=30)
                payload = _parse_json_payload(r.text)
            except Exception:
                payload = None

            data = (payload or {}).get("data") or []
            if not data:
                if page > 1:
                    break
                continue

            for it in data:
                name = _normalize_model(str(it.get("product_name") or ""))
                slug = str(it.get("slug") or "")
                if not name:
                    continue
                # Strict brand filter to avoid cross-brand bikes.
                if brand_l not in name.lower() and brand_l not in slug.lower():
                    continue
                if not _looks_like_cycle_model(name):
                    continue
                key = name.lower()
                if key in seen:
                    continue
                seen.add(key)

                listed = _parse_price_num(it.get("product_price") or it.get("mrp"))
                disc = _parse_price_num(it.get("special_price") or it.get("sales_price") or it.get("price"))
                if listed is not None:
                    listed = _to_inr(listed, "INR")
                if disc is not None:
                    disc = _to_inr(disc, "INR")
                if listed and disc and listed < disc:
                    listed = None

                segment = sub.replace("-", " ").title()
                rows.append(_row(name, segment, listed, disc))

    return _dedupe_rows(rows)


def scrape_sportnetwork_slug_crawl(brand_name: str):
    """Crawl SportNetwork brand pages for resolved product slugs using Playwright.

    This captures `brand-product-page/<brand>/<slug>` URLs from both DOM links and
    network responses while scrolling/clicking load-more style controls.
    """
    rows = []
    brand_l = (brand_name or "").strip().lower()
    if brand_l not in {"scott", "avanti", "bergamont"}:
        return rows

    try:
        from playwright.sync_api import sync_playwright
    except Exception:
        return rows

    # Pull price hints from API rows so slug-crawled models can inherit INR prices.
    api_rows = scrape_sportnetwork_api_brand(brand_name)
    price_map = {}
    for r in api_rows:
        model = _normalize_model(r.get("Model name", ""))
        if not model:
            continue
        price_map[model.lower()] = (
            r.get("Listed price", ""),
            r.get("Discounted price", ""),
            r.get("Segment", ""),
        )

    slug_set = set()
    seed_urls = _sportnetwork_brand_seed_urls(brand_name)

    def _add_slug_from_url(url: str):
        m = re.search(r"/brand-product-page/([a-z0-9-]+)/([^?&#]+)", url, flags=re.I)
        if not m:
            return
        b = (m.group(1) or "").lower()
        slug = (m.group(2) or "").strip("/")
        if b != brand_l or not slug:
            return
        slug_set.add(slug)

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=True)
        page = browser.new_page(extra_http_headers={"Accept-Language": "en-IN,en;q=0.9"})
        started = time.monotonic()
        time_budget = 240 if brand_l == "avanti" else 165

        def _time_exceeded():
            return (time.monotonic() - started) > time_budget

        def on_response(resp):
            u = resp.url
            if "/brand-product-page/" in u:
                _add_slug_from_url(u)
            if "sportnetwork.in" not in u:
                return
            if not any(k in u for k in ["dc.sportnetwork.in", "/_next/data/", "/brand-products-list/"]):
                return
            try:
                body = resp.text()
            except Exception:
                body = ""
            for s in _extract_brand_product_slugs(body[:350000], brand_l):
                slug_set.add(s)

        page.on("response", on_response)

        queue = list(seed_urls)
        seen_pages = set()
        max_page_visits = 80 if brand_l == "avanti" else 55

        while queue and len(seen_pages) < max_page_visits:
            if _time_exceeded():
                break
            u = queue.pop(0)
            if not u or u in seen_pages:
                continue
            if "sportnetwork.in" not in u:
                continue
            seen_pages.add(u)

            try:
                page.goto(u, wait_until="domcontentloaded", timeout=60000)
                page.wait_for_timeout(1800)
            except Exception:
                continue

            # Scroll deep to trigger lazy lists.
            for _ in range(20 if brand_l == "avanti" else 14):
                if _time_exceeded():
                    break
                page.mouse.wheel(0, 7000)
                page.wait_for_timeout(420)

            # Aggressive pagination controls for Avanti and siblings.
            click_labels = [
                "Load More", "Show More", "View More", "Next", "Next Page",
                "Older", ">", "»", "MORE"
            ]
            for _ in range(16 if brand_l == "avanti" else 10):
                if _time_exceeded():
                    break
                clicked = False
                for lbl in click_labels:
                    try:
                        btn = page.get_by_text(lbl, exact=False).first
                        if btn and btn.is_visible():
                            btn.click(timeout=1600)
                            page.wait_for_timeout(1000)
                            clicked = True
                    except Exception:
                        pass
                if not clicked:
                    break

            # Capture links from DOM and queue more catalog pages recursively.
            try:
                links = page.eval_on_selector_all("a[href]", "els => els.map(e => e.getAttribute('href'))")
            except Exception:
                links = []
            for h in links or []:
                if not h:
                    continue
                full = urljoin("https://sportnetwork.in", h)
                lu = full.lower()
                if "/brand-product-page/" in lu:
                    _add_slug_from_url(full)
                    continue
                if "/products-list/bike/" in lu or "/brand-products-list/" in lu:
                    if not any(x in lu for x in ["/equipments", "/apparel", "/accessories", "/book-your-service"]):
                        if full not in seen_pages and full not in queue:
                            queue.append(full)
                # Numeric pagination links.
                if "page=" in lu and ("products-list" in lu or "brand-products-list" in lu):
                    if full not in seen_pages and full not in queue:
                        queue.append(full)

            # Parse current HTML for hidden slug references.
            try:
                html_blob = page.content()
            except Exception:
                html_blob = ""
            for s in _extract_brand_product_slugs(html_blob[:500000], brand_l):
                slug_set.add(s)

        # Recursive expansion from discovered product pages.
        product_queue = list(slug_set)
        seen_product_pages = set()
        max_product_visits = 520 if brand_l == "avanti" else 320
        while product_queue and len(seen_product_pages) < max_product_visits:
            if _time_exceeded():
                break
            slug = product_queue.pop(0)
            if not slug or slug in seen_product_pages:
                continue
            seen_product_pages.add(slug)
            purl = f"https://sportnetwork.in/brand-product-page/{brand_l}/{slug}"
            try:
                page.goto(purl, wait_until="domcontentloaded", timeout=45000)
                page.wait_for_timeout(900)
                html_blob = page.content()
            except Exception:
                continue

            # Collect sibling product slugs from recommendation blocks / embedded payloads.
            for s in _extract_brand_product_slugs(html_blob[:700000], brand_l):
                if s not in slug_set:
                    slug_set.add(s)
                    if s not in seen_product_pages:
                        product_queue.append(s)

            try:
                links = page.eval_on_selector_all("a[href]", "els => els.map(e => e.getAttribute('href'))")
            except Exception:
                links = []
            for h in links or []:
                if not h:
                    continue
                full = urljoin("https://sportnetwork.in", h)
                lu = full.lower()
                if "/brand-product-page/" in lu:
                    _add_slug_from_url(full)
                    m = re.search(r"/brand-product-page/[a-z0-9-]+/([^?&#/]+)", full, flags=re.I)
                    if m:
                        s2 = m.group(1).strip("/")
                        if s2 and s2 not in seen_product_pages:
                            product_queue.append(s2)

        browser.close()

    # Convert slugs to rows and enrich price/segment from API map if present.
    for slug in sorted(slug_set):
        model = _normalize_model(slug.replace("-", " "))
        if not _looks_like_cycle_model(model):
            continue

        listed = ""
        disc = ""
        seg = ""
        p = price_map.get(model.lower())
        if p:
            listed, disc, seg = p

        # Derive a segment hint from model text when API segment is missing.
        lm = model.lower()
        if not seg:
            if any(k in lm for k in ["mountain", "mtb", "scale", "spark", "genius", "ransom", "aspect"]):
                seg = "Mountain"
            elif any(k in lm for k in ["road", "addict", "foil", "rc"]):
                seg = "Road"
            elif any(k in lm for k in ["gravel", "cyclocross", "speedster gravel"]):
                seg = "Gravel"
            elif any(k in lm for k in ["urban", "city", "sub", "hybrid"]):
                seg = "Hybrid/Urban"
            elif any(k in lm for k in ["e-bike", "ebike", "e bike"]):
                seg = "E-Bike"

        rows.append(_row(model, seg, listed, disc))

    return _dedupe_rows(rows)


def scrape_basso_links(cfg):
    rows = []
    seen = set()
    try:
        html = _safe_get("https://bassobikes.com/en").text
    except Exception:
        return rows

    hrefs = re.findall(r'href="([^\"]+)"', html, re.I)
    for h in hrefs:
        full = urljoin("https://bassobikes.com", h)
        lu = full.lower()
        if "/en/bikes/" not in lu:
            continue
        if any(k in lu for k in ["/configurator", "/stories/", "/news/", "/blog/"]):
            continue
        model = _model_from_url(full)
        if not _looks_like_cycle_model(model):
            continue
        key = model.lower()
        if key in seen:
            continue
        seen.add(key)

        seg = ""
        if "/road-bikes/" in lu:
            seg = "Road"
        elif "/gravel-bikes/" in lu:
            seg = "Gravel"
        elif "/e-bikes/" in lu:
            seg = "E-Bike"
        rows.append(_row(model, seg, "", ""))
    return rows


def apply_strict_phase2(cfg, rows):
    brand_l = cfg["name"].lower()
    if brand_l not in STRICT_PHASE2_BRANDS:
        return rows

    clean_count = len([r for r in rows if not str(r.get("Model name", "")).startswith("[")])
    if clean_count >= 80:
        return rows

    extras = []
    if brand_l in {"scott", "avanti", "bergamont"}:
        extras.extend(scrape_sportnetwork_api_brand(cfg["name"]))
        extras.extend(scrape_sportnetwork_slug_crawl(cfg["name"]))
    if brand_l == "merida":
        extras.extend(scrape_merida_bikefinder(cfg))
    if brand_l == "cube":
        extras.extend(scrape_cube_sitemap(cfg))
    if brand_l == "java":
        extras.extend(scrape_java_sitemap(cfg))
    if brand_l == "jamis":
        extras.extend(scrape_jamis_sitemap(cfg))
    if brand_l == "orbea":
        extras.extend(scrape_orbea_playwright(cfg))
    if brand_l == "basso":
        extras.extend(scrape_basso_links(cfg))
        extras.extend(scrape_html_jsonld(cfg))

    if extras:
        rows = _dedupe_rows(list(rows) + list(extras))
    return rows


def apply_global_deepening(cfg, rows):
    """Deepen crawl for all brands by adding sitemap/domain enrichment when counts are low."""
    clean_count = len([r for r in rows if not str(r.get("Model name", "")).startswith("[")])
    if clean_count >= 180:
        return rows

    extras = []
    method = cfg.get("method", "")

    # Always try sitemap enrichment when we are below expected range.
    if clean_count < 140:
        try:
            extras.extend(scrape_sitemap_models(cfg["url"], cfg["name"]))
        except Exception:
            pass

    # For generic/html brands, do another deeper page graph crawl.
    if method == "html_jsonld" and clean_count < 160:
        try:
            extras.extend(scrape_html_jsonld(cfg))
        except Exception:
            pass

    # For non-HTML methods that still underperform, backfill from HTML parser once.
    if method in {"trek_occ", "canyon_json", "shopify", "sportnetwork_brand"} and clean_count < 90:
        try:
            extras.extend(scrape_html_jsonld(cfg))
        except Exception:
            pass

    # Extra Avanti hardening pass.
    if cfg["name"].lower() == "avanti" and clean_count < 80:
        try:
            extras.extend(scrape_sportnetwork_slug_crawl("Avanti"))
        except Exception:
            pass

    if extras:
        rows = _dedupe_rows(list(rows) + list(extras))
    return rows


def _dedupe_rows(rows):
    best = {}
    for r in rows:
        model = _normalize_model(r.get("Model name", ""))
        if not _looks_like_cycle_model(model):
            continue
        seg = (r.get("Segment", "") or "").strip()
        listed = _parse_price_num(r.get("Listed price"))
        disc = _parse_price_num(r.get("Discounted price"))
        key = model.lower()
        prev = best.get(key)
        if prev is None:
            best[key] = _row(model, seg, listed, disc)
            continue
        # Keep richer segment and better prices.
        prev_seg = prev.get("Segment", "")
        if not prev_seg and seg:
            prev["Segment"] = seg
        prev_l = _parse_price_num(prev.get("Listed price"))
        prev_d = _parse_price_num(prev.get("Discounted price"))
        if listed and (not prev_l or listed > prev_l):
            prev["Listed price"] = listed
        if disc and (not prev_d or disc < prev_d):
            prev["Discounted price"] = disc
    out = list(best.values())
    out.sort(key=lambda x: x["Model name"].lower())
    return out


def scrape_trek_occ(cfg):
    rows = []
    base = "https://api.trekbikes.com"
    site = "https://www.trekbikes.com"
    store = "in"
    lang = "en_IN"
    # Category B100 is bikes, but let's just make sure we hit the main bikes category
    categories = ["10000", "12199", "12140", "11652", "11651", "12470", "11650"]
    
    for category in categories:
        page = 0
        max_pages = 8
        while page < max_pages:
            url = (
                f"{base}/occ/v2/{store}/products/search?fields=FULL"
                f"&query=%3Arelevance%3Acategory%3A{category}"
                f"&lang={lang}&curr=INR&pageSize=100&currentPage={page}"
            )
            try:
                r = requests.get(url, headers={**HDR, "Origin": site, "Referer": site + "/"}, timeout=20)
                r.raise_for_status()
                data = r.json()
            except Exception:
                break
            prods = data.get("products", [])
            if not prods:
                break
            for p in prods:
                # Need to look deeper into the specific structure of OCC
                raw_name = p.get("name", "")
                # Skip things that are obviously components/gear
                if any(kw in raw_name.lower() for kw in ["kit", "pedal", "light", "bottle", "helmet", "shoe", "tyre", "tire", "saddle", "glove", "bag"]):
                    continue
                
                model = _normalize_model(raw_name)
                if not _looks_like_cycle_model(model):
                    continue
                segment = str(p.get("defaultCategory", "") or "").title()
                
                # OCC Prices are sometimes inside `price` directly, or `offerPrice`
                price_obj = p.get("price") or {}
                if not price_obj:
                    price_obj = p.get("stock", {}).get("price") or {}
                
                current_raw = price_obj.get("value")
                if current_raw is None and price_obj.get("formattedValue"):
                    current_raw = _parse_price_num(price_obj.get("formattedValue"))
                else:
                    current_raw = _parse_price_num(current_raw)
                    
                listed_raw = (p.get("wasPrice") or {}).get("value")
                listed_raw = _parse_price_num(listed_raw)
                
                if current_raw:
                    if listed_raw and listed_raw > current_raw:
                        rows.append(_row(model, segment, listed_raw, current_raw))
                    else:
                        rows.append(_row(model, segment, current_raw, current_raw))
                else:
                    # If we don't have a current price, try basePrice or some other field
                    base_price = (p.get("basePrice") or {}).get("value")
                    base_price = _parse_price_num(base_price)
                    if base_price:
                        rows.append(_row(model, segment, base_price, base_price))
                    elif listed_raw:
                         rows.append(_row(model, segment, listed_raw, listed_raw))
            
            pagination = data.get("pagination", {})
            if page + 1 >= pagination.get("totalPages", 1):
                break
            page += 1

    if rows:
        return _dedupe_rows(rows)

    # Fallback from sitemap if OCC returns empty.
    rows.extend(scrape_sitemap_models("https://www.trekbbikes.com/in/en_IN", "Trek"))
    if len(rows) < 20:
        rows.extend(scrape_sitemap_models("https://www.trekbikes.com/us/en_US", "Trek"))
    return _dedupe_rows(rows)


def scrape_shopify(cfg):
    url_base = cfg["url"].rstrip("/")
    brand = cfg["name"].lower()
    cur = SHOPIFY_CURRENCY_HINT.get(brand, "USD")
    rows = []
    page = 1
    while True:
        url = f"{url_base}/products.json?limit=250&page={page}"
        try:
            data = _safe_get(url).json()
        except Exception:
            break
        products = data.get("products", [])
        if not products:
            break
        for p in products:
            model = _normalize_model(p.get("title", ""))
            if not _looks_like_cycle_model(model):
                continue
            seg = (p.get("product_type") or "").strip()
            listed_vals, disc_vals = [], []
            for v in p.get("variants", []):
                pv = _parse_price_num(v.get("price"))
                cv = _parse_price_num(v.get("compare_at_price"))
                if pv:
                    disc_vals.append(_to_inr(pv, cur))
                if cv:
                    listed_vals.append(_to_inr(cv, cur))
            listed = max(listed_vals) if listed_vals else None
            disc = min(disc_vals) if disc_vals else None
            if listed and disc and listed < disc:
                listed = None
            rows.append(_row(model, seg, listed, disc))
        page += 1
        time.sleep(0.2)
    return _dedupe_rows(rows)


def scrape_canyon_json(cfg):
    base = "https://www.canyon.com/en-in"
    paths = [
        "/road-bikes/",
        "/mountain-bikes/",
        "/hybrid-bikes/",
        "/gravel-bikes/",
        "/electric-bikes/",
    ]
    rows = []
    seen_links = set()

    for p in paths:
        try:
            html = _safe_get(base + p).text
        except Exception:
            continue
        soup = BeautifulSoup(html, "html.parser")

        # JSON-LD products.
        for s in soup.select('script[type="application/ld+json"]'):
            raw = s.string or s.get_text("", strip=True)
            if not raw:
                continue
            try:
                obj = json.loads(raw)
            except Exception:
                continue
            payloads = obj if isinstance(obj, list) else [obj]
            for it in payloads:
                if not isinstance(it, dict):
                    continue
                if it.get("@type") not in {"Product", "ListItem", "ItemList"}:
                    continue
                if it.get("@type") == "Product":
                    model = _normalize_model(it.get("name", ""))
                    if not _looks_like_cycle_model(model):
                        continue
                    cat = it.get("category", "")
                    offers = it.get("offers") or {}
                    if isinstance(offers, list):
                        offers = offers[0] if offers else {}
                    price = _to_inr(_parse_price_num(offers.get("price")), _detect_currency(str(offers), _host(base)))
                    rows.append(_row(model, cat, "", price))

        # HTML card fallback.
        cards = soup.select("a[href*='/bikes/'], a[href*='/road-bikes/'], a[href*='/mountain-bikes/']")
        for a in cards:
            href = a.get("href", "")
            if not href:
                continue
            full = urljoin(base, href)
            if full in seen_links:
                continue
            seen_links.add(full)
            t = a.get_text(" ", strip=True)
            model = _normalize_model(t)
            if not _looks_like_cycle_model(model):
                continue
            card_txt = " ".join(a.parent.stripped_strings) if a.parent else t
            curr = _detect_currency(card_txt, _host(base))
            disc = _to_inr(_parse_price_num(card_txt), curr)
            rows.append(_row(model, "", "", disc))

    return _dedupe_rows(rows)


def _sportnetwork_brand_url(brand):
    b = brand.lower()
    if b == "scott":
        return "https://sportnetwork.in/brand-products-list/scott"
    if b == "avanti":
        return "https://sportnetwork.in/brand-products-list/Avanti"
    if b == "bergamont":
        return "https://sportnetwork.in/brand-products-list/bergamont"
    return "https://sportnetwork.in/"


def _walk_json_for_products(obj, out):
    if isinstance(obj, dict):
        keys = {k.lower(): k for k in obj.keys()}
        name_k = keys.get("name") or keys.get("title") or keys.get("productname")
        price_k = keys.get("price") or keys.get("saleprice") or keys.get("regularprice")
        seg_k = keys.get("category") or keys.get("segment")
        if name_k and price_k:
            out.append({
                "name": obj.get(name_k),
                "segment": obj.get(seg_k, "") if seg_k else "",
                "price": obj.get(price_k),
                "raw": str(obj),
            })
        for v in obj.values():
            _walk_json_for_products(v, out)
    elif isinstance(obj, list):
        for it in obj:
            _walk_json_for_products(it, out)


def scrape_sportnetwork_brand(cfg):
    url = _sportnetwork_brand_url(cfg["name"])
    rows = []
    try:
        html = _safe_get(url).text
    except Exception:
        return rows

    soup = BeautifulSoup(html, "html.parser")
    nd = soup.find("script", id="__NEXT_DATA__")
    if nd and nd.string:
        try:
            data = json.loads(nd.string)
            hits = []
            _walk_json_for_products(data, hits)
            for h in hits:
                model = _normalize_model(str(h.get("name", "")))
                if not _looks_like_cycle_model(model):
                    continue
                raw = h.get("raw", "")
                curr = _detect_currency(raw, "sportnetwork.in")
                disc = _to_inr(_parse_price_num(h.get("price")), curr)
                rows.append(_row(model, str(h.get("segment", "")), "", disc))
        except Exception:
            pass

    # Fallback selectors if NEXT payload shape changes.
    if not rows:
        for card in soup.select("article, li, div"):
            txt = card.get_text(" ", strip=True)
            if len(txt) < 20:
                continue
            if "₹" not in txt and "$" not in txt and "price" not in txt.lower():
                continue
            a = card.select_one("a[href]")
            if not a:
                continue
            model = _normalize_model(a.get_text(" ", strip=True))
            if not _looks_like_cycle_model(model):
                continue
            curr = _detect_currency(txt, "sportnetwork.in")
            disc = _to_inr(_parse_price_num(txt), curr)
            rows.append(_row(model, "", "", disc))

    if not rows:
        rows = scrape_sitemap_models(cfg["url"], cfg["name"])
    return _dedupe_rows(rows)


def scrape_html_jsonld(cfg):
    base = cfg["url"].rstrip("/")
    rows = []
    seed_pages = [base]

    # Common catalog entry paths.
    for p in [
        "/bikes", "/bikes/", "/collections/all-bikes", "/road-bikes", "/mountain-bikes",
        "/gravel-bikes", "/hybrid-bikes", "/electric-bikes", "/products", "/collections/bikes",
        "/catalog", "/shop/bikes"
    ]:
        seed_pages.append(base + p)

    allow_kw = ["bike", "bikes", "cycle", "road", "mountain", "gravel", "mtb", "hybrid", "e-bike", "ebike"]
    deny_kw = ["helmet", "jersey", "shoe", "glove", "socks", "accessories", "apparel", "parts", "news", "blog", "story", "faq", "dealer", "warranty", "service"]
    base_host = _host(base)

    # Domain BFS for deeper crawling.
    queue = [(u, 0) for u in seed_pages]
    seen_pages = set()
    max_pages = 65
    max_depth = 2

    while queue and len(seen_pages) < max_pages:
        u, depth = queue.pop(0)
        if not u or u in seen_pages:
            continue
        if _host(u) != base_host:
            continue
        seen_pages.add(u)
        try:
            html = _safe_get(u).text
        except Exception:
            continue
        soup = BeautifulSoup(html, "html.parser")

        # JSON-LD extraction.
        for s in soup.select('script[type="application/ld+json"]'):
            raw = s.string or s.get_text("", strip=True)
            if not raw:
                continue
            try:
                obj = json.loads(raw)
            except Exception:
                continue
            stack = obj if isinstance(obj, list) else [obj]
            while stack:
                it = stack.pop()
                if isinstance(it, list):
                    stack.extend(it)
                    continue
                if not isinstance(it, dict):
                    continue
                t = it.get("@type", "")
                if t == "Product":
                    model = _normalize_model(it.get("name", ""))
                    if not _looks_like_cycle_model(model):
                        continue
                    seg = it.get("category", "")
                    offers = it.get("offers") or {}
                    if isinstance(offers, list):
                        offers = offers[0] if offers else {}
                    raw_offer = str(offers)
                    cur = _detect_currency(raw_offer or raw, _host(base))
                    raw_price = offers.get("price") or offers.get("lowPrice")
                    raw_listed = offers.get("highPrice") or offers.get("priceBeforeDiscount") or offers.get("price")
                    disc = _to_inr(_parse_price_num(raw_price), cur) if raw_price is not None else None
                    listed = _to_inr(_parse_price_num(raw_listed), cur) if raw_listed is not None else None
                    if listed and disc and listed <= disc:
                        listed = None
                    if disc is not None:
                        rows.append(_row(model, seg, listed, disc))
                for v in it.values():
                    if isinstance(v, (dict, list)):
                        stack.append(v)

        # HTML product card fallback.
        for card in soup.select("article, li, div"):
            a = card.select_one("a[href]")
            if not a:
                continue
            href = a.get("href", "")
            if not href or href.startswith("#"):
                continue
            text = card.get_text(" ", strip=True)
            if len(text) < 15:
                continue
            
            # Identify a specific price element if possible
            price_el = card.select_one(".product-price, [data-price], .price, .money, .sales-price, .regular-price, [data-product-price]")
            found_explicit_el = False
            if price_el:
                price_text = price_el.get_text(" ", strip=True) + " " + (price_el.get("content") or "") + " " + (price_el.get("data-price") or "")
                found_explicit_el = True
            else:
                price_text = text

            if not re.search(r"(₹|\$|€|£|₱|INR|USD|EUR|GBP|PHP|AED|price)", price_text, re.I):
                # Fallback to general text if price element is generic and text has symbols
                if not re.search(r"(₹|\$|€|£|₱|INR|USD|EUR|GBP|PHP|AED)", text, re.I):
                    continue

            model = _normalize_model(a.get_text(" ", strip=True))
            if not _looks_like_cycle_model(model):
                continue
            cur = _detect_currency(price_text or text, _host(base))
            
            # Try to grab explicit sale and list prices
            # Prioritize numbers with explicit currency symbols
            sym_pattern = r"(?:₹|\$|€|£|₱|INR|USD|EUR|GBP|PHP|AED)\s*([\d][\d,]*(?:\.\d{1,2})?)"
            num_pattern = r"\b([\d][\d,]*(?:\.\d{1,2})?)\b"
            
            raw_prices_sym = re.findall(sym_pattern, price_text, re.I)
            
            if raw_prices_sym:
                prices = raw_prices_sym
            else:
                # If no currency symbol found, only accept if we found an explicit price element, 
                # or if the number is large enough to be a price (e.g. > 100 for USD/EUR, > 5000 for INR), avoiding models/years
                prices = []
                for p in re.findall(num_pattern, price_text):
                    val = float(p.replace(",", ""))
                    if val == 2023 or val == 2024 or val == 2025: continue
                    # Filter out small integers from model names like 105, 5, 500 unless wrapped in explicit price HTML
                    if found_explicit_el or (cur == 'INR' and val > 5000) or (cur != 'INR' and val > 250):
                        prices.append(p)

            if len(prices) >= 2:
                disc = _to_inr(_parse_price_num(prices[0]), cur)
                listed = _to_inr(_parse_price_num(prices[1]), cur)
                if listed and disc and listed < disc:
                    listed, disc = disc, listed
            elif len(prices) == 1:
                disc = _to_inr(_parse_price_num(prices[0]), cur)
                listed = None
            else:
                disc, listed = None, None
                listed = None

            rows.append(_row(model, "", listed, disc))

        # Expand deeper links within the same host for bike-like routes.
        if depth < max_depth:
            for a in soup.select("a[href]"):
                href = a.get("href", "")
                if not href or href.startswith("#"):
                    continue
                full = urljoin(base + "/", href)
                lu = full.lower()
                if _host(full) != base_host:
                    continue
                if any(k in lu for k in deny_kw):
                    continue
                if not any(k in lu for k in allow_kw):
                    continue
                if full not in seen_pages:
                    queue.append((full, depth + 1))

    # Always enrich with sitemap links if row volume is still low.
    if len(rows) < 140:
        rows.extend(scrape_sitemap_models(cfg["url"], cfg["name"]))
    return _dedupe_rows(rows)


def dispatch_scrape(cfg):
    m = cfg["method"]
    if m == "trek_occ":
        return scrape_trek_occ(cfg)
    if m == "shopify":
        return scrape_shopify(cfg)
    if m == "canyon_json":
        return scrape_canyon_json(cfg)
    if m == "sportnetwork_brand":
        return scrape_sportnetwork_brand(cfg)
    if m == "html_jsonld":
        return scrape_html_jsonld(cfg)
    return []


def sync_database_and_get_deltas(brand_rows, db_path="prices_history.db"):
    conn = sqlite3.connect(db_path)
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS prices (
            brand TEXT,
            model_name TEXT,
            segment TEXT,
            listed_price REAL,
            discounted_price REAL,
            last_seen TIMESTAMP,
            first_seen TIMESTAMP,
            PRIMARY KEY (brand, model_name)
        )
    """)
    conn.commit()

    deltas = []
    now = datetime.now()

    for brand, rows in brand_rows:
        for r in rows:
            model = r.get("Model name")
            if not model or model == "[No products scraped]":
                continue
            seg = r.get("Segment", "")
            listed = r.get("Listed price") or 0.0
            disc = r.get("Discounted price") or 0.0

            c.execute("SELECT discounted_price FROM prices WHERE brand=? AND model_name=?", (brand, model))
            existing = c.fetchone()

            if existing:
                old_price = existing[0]
                if old_price and disc and disc < old_price:
                    drop_pct = round(((old_price - disc) / old_price) * 100, 1)
                    deltas.append({
                        "Brand": brand,
                        "Model name": model,
                        "Old Price": old_price,
                        "New Price": disc,
                        "Drop %": f"{drop_pct}%"
                    })
                c.execute("""
                    UPDATE prices 
                    SET segment=?, listed_price=?, discounted_price=?, last_seen=?
                    WHERE brand=? AND model_name=?
                """, (seg, listed, disc, now, brand, model))
            else:
                c.execute("""
                    INSERT INTO prices (brand, model_name, segment, listed_price, discounted_price, first_seen, last_seen)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (brand, model, seg, listed, disc, now, now))
    
    conn.commit()
    conn.close()
    return deltas

def save_xlsx(brand_rows, out_path=OUT_XLSX):
    deltas = sync_database_and_get_deltas(brand_rows)
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    summary = wb.create_sheet("Summary")
    summary.append(["Brand", "Rows"])
    for c in range(1, 3):
        cell = summary.cell(1, c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F81BD")
        cell.alignment = Alignment(horizontal="center")

    alerts = wb.create_sheet("Price Alerts")
    alerts.append(["Brand", "Model name", "Old Price", "New Price", "Drop %"])
    for c in range(1, 6):
        cell = alerts.cell(1, c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="C0504D")
        cell.alignment = Alignment(horizontal="center")
    for d in deltas:
        alerts.append([d["Brand"], d["Model name"], d["Old Price"], d["New Price"], d["Drop %"]])

    for brand, rows in brand_rows:
        title = re.sub(r"[\\[\\]:*?/]", "-", brand)[:31] or "Sheet"
        ws = wb.create_sheet(title)
        ws.append(COLS)
        for c in range(1, len(COLS) + 1):
            cell = ws.cell(1, c)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4F81BD")
            cell.alignment = Alignment(horizontal="center")

        for r in rows:
            ws.append([r.get(c, "") for c in COLS])

        for col_i in range(1, len(COLS) + 1):
            letter = get_column_letter(col_i)
            max_len = len(COLS[col_i - 1])
            for rr in ws.iter_rows(min_row=2, min_col=col_i, max_col=col_i):
                v = rr[0].value
                if v is None:
                    continue
                max_len = max(max_len, len(str(v)))
            ws.column_dimensions[letter].width = min(max(14, max_len + 2), 60)

        summary.append([brand, len(rows)])

    for c in range(1, 3):
        letter = get_column_letter(c)
        summary.column_dimensions[letter].width = 28 if c == 1 else 12

    for c in range(1, 6):
        letter = get_column_letter(c)
        alerts.column_dimensions[letter].width = 25

    wb.save(out_path)


def process_brand(cfg):
    print(f"\n[START] {cfg['name']} [{cfg['method']}]")
    try:
        rows = dispatch_scrape(cfg)
    except Exception as e:
        print(f"[{cfg['name']}] error: {e}")
        rows = []
    try:
        rows = apply_strict_phase2(cfg, rows)
    except Exception as e:
        print(f"[{cfg['name']}] strict phase2 warning: {e}")
    try:
        rows = apply_global_deepening(cfg, rows)
    except Exception as e:
        print(f"[{cfg['name']}] global deepen warning: {e}")

    # --- ADVANCED DDG FALLBACK LAYER ---
    missing_count = 0
    max_lookups_per_brand = 15 # Cap to prevent DDG IP Ban
    for r in (rows or []):
        if r.get("Model name") and r.get("Model name") != "[No products scraped]":
            if not r.get("Discounted price"):
                if missing_count < max_lookups_per_brand:
                    try:
                        time.sleep(1 + random.random()*1.5) # Evasion delay
                        search_p = fetch_price_from_search(cfg["name"], r["Model name"])
                        if search_p:
                            r["Discounted price"] = search_p
                            r["Listed price"] = r.get("Listed price") or search_p
                            print(f"[{cfg['name']}] 🔍 Filled via Search: {r['Model name']} -> ₹{search_p}")
                    except Exception:
                        pass
                    missing_count += 1

    if not rows:
        rows = [_row("[No products scraped]", "", "", "")]
    print(f"[DONE] {cfg['name']} - rows: {len(rows)}")
    return cfg["name"], rows

def main():
    args = [a.strip() for a in sys.argv[1:] if a.strip()]
    out_file = OUT_XLSX
    
    if "--out" in args:
        idx = args.index("--out")
        out_file = args[idx + 1]
        args.pop(idx)
        args.pop(idx)
        
    filters = {a.lower() for a in args if a and not a.startswith("--")}

    print("=" * 70)
    print("Global Brands Scraper (Multi-threaded)")
    print("=" * 70)

    target_brands = [cfg for cfg in BRANDS if not filters or cfg["name"].lower() in filters]
    results = []
    
    with ThreadPoolExecutor(max_workers=5) as executor:
        futures = {executor.submit(process_brand, cfg): cfg for cfg in target_brands}
        for future in as_completed(futures):
            brand_name, rows = future.result()
            results.append((brand_name, rows))

    # Optional: Sort results alphabetically by brand name for a consistent Excel sheet order
    results.sort(key=lambda x: x[0].lower())
    
    save_xlsx(results, out_file)
    print(f"\nSaved -> {out_file}")


if __name__ == "__main__":
    main()
